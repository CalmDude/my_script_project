"""
S&P 500 Scanner - Find stocks with FULL HOLD + ADD signal
Scans all S&P 500 stocks for strongest bullish alignment (weekly P1 + daily P1)
Only outputs stocks with "FULL HOLD + ADD" signal
"""

import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from technical_analysis import analyze_ticker
from datetime import datetime
import yfinance as yf
from pathlib import Path
import logging

# Configure logging
logger = logging.getLogger(__name__)
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def safe_print(text):
    """Print text with fallback for Windows console encoding issues"""
    try:
        print(text)
    except UnicodeEncodeError:
        # Remove emojis and special characters for Windows console
        import re

        # Remove emoji patterns
        clean_text = re.sub(
            r"[\U0001F300-\U0001F9FF\u2600-\u26FF\u2700-\u27BF]", "", text
        )
        print(clean_text)


def determine_regime_tier(signal, weekly_state, daily_state):
    """
    Determine market regime tier based on QQQ Larsson signal.

    Tier Logic:
    - GREEN (FULL HOLD + ADD): Trade normally, EXCELLENT/GOOD quality
    - YELLOW (HOLD, SCALE IN): Ultra-selective, EXCELLENT + SAFE ENTRY + 3:1+ Vol R:R only
    - ORANGE (HOLD MOST + REDUCE, LIGHT/CASH): No new buys, portfolio exits only
    - RED (REDUCE, CASH, FULL CASH/DEFEND): No new buys, portfolio exits only

    Args:
        signal: QQQ Larsson signal
        weekly_state: Weekly trend state (P1/P2/N1/N2)
        daily_state: Daily trend state (P1/P2/N1/N2)

    Returns:
        dict with tier, color, emoji, message, allow_buys
    """
    signal_upper = signal.upper()

    # GREEN LIGHT: Full bullish alignment
    if signal_upper == "FULL HOLD + ADD":
        return {
            "tier": "GREEN",
            "color": "green",
            "emoji": "🟢",
            "message": "FULL BULL MARKET - Trade normally",
            "allow_buys": True,
            "filter_mode": "normal",  # EXCELLENT, GOOD quality
        }

    # YELLOW LIGHT: Constructive but selective
    elif signal_upper in ["HOLD", "SCALE IN"]:
        return {
            "tier": "YELLOW",
            "color": "yellow",
            "emoji": "🟡",
            "message": "SELECTIVE MARKET - EXCELLENT + SAFE ENTRY + 3:1+ Vol R:R only",
            "allow_buys": True,
            "filter_mode": "strict",  # EXCELLENT + SAFE ENTRY + 3:1+ only
        }

    # ORANGE LIGHT: Deteriorating but not full defense
    elif signal_upper in ["HOLD MOST + REDUCE", "LIGHT / CASH"]:
        return {
            "tier": "ORANGE",
            "color": "orange",
            "emoji": "🟠",
            "message": "RISK OFF - No new buys, exits only",
            "allow_buys": False,
            "filter_mode": "exits_only",
        }

    # RED LIGHT: Full defensive mode
    elif signal_upper in ["REDUCE", "CASH", "FULL CASH / DEFEND"]:
        return {
            "tier": "RED",
            "color": "red",
            "emoji": "🔴",
            "message": "BEAR MARKET - No new buys, preserve capital",
            "allow_buys": False,
            "filter_mode": "exits_only",
        }

    # Unknown signal - default to ORANGE (cautious)
    else:
        return {
            "tier": "ORANGE",
            "color": "orange",
            "emoji": "⚠️",
            "message": f"UNKNOWN SIGNAL: {signal} - Defaulting to cautious mode",
            "allow_buys": False,
            "filter_mode": "exits_only",
        }


def analyze_market_regime(daily_bars=60, weekly_bars=52, as_of_date=None):
    """
    Analyze QQQ (NASDAQ 100 ETF) to determine market regime.

    This is MANDATORY and runs before all stock scanning.
    The regime determines what trades are allowed:
    - GREEN: Normal trading
    - YELLOW: Ultra-selective (EXCELLENT + SAFE ENTRY + 3:1+ only)
    - ORANGE/RED: No new buys, exits only

    Args:
        daily_bars: Number of daily bars for analysis
        weekly_bars: Number of weekly bars for analysis
        as_of_date: Optional historical date for backtesting

    Returns:
        dict with regime info (tier, signal, states, emoji, message, etc.)
    """
    print("\n" + "=" * 80)
    print("MARKET REGIME ANALYSIS (QQQ - NASDAQ 100 ETF)")
    print("=" * 80)

    try:
        # Analyze QQQ using same technical analysis
        qqq_result = analyze_ticker("QQQ", daily_bars, weekly_bars, as_of_date)

        if "error" in qqq_result:
            print(f"⚠️  WARNING: Could not analyze QQQ: {qqq_result['error']}")
            print("⚠️  DEFAULTING TO ORANGE (CAUTIOUS) - No new buys allowed")
            return {
                "tier": "ORANGE",
                "signal": "ERROR",
                "weekly_state": "N/A",
                "daily_state": "N/A",
                "color": "orange",
                "emoji": "⚠️",
                "message": "QQQ analysis failed - Defaulting to cautious mode",
                "allow_buys": False,
                "filter_mode": "exits_only",
                "qqq_price": None,
                "qqq_rsi": None,
            }

        # Extract signal and states
        signal = qqq_result.get("signal", "UNKNOWN")
        weekly_state = qqq_result.get("weekly_state", "N/A")
        daily_state = qqq_result.get("daily_state", "N/A")
        qqq_price = qqq_result.get("current_price", None)
        qqq_rsi = qqq_result.get("rsi", None)

        # Determine regime tier
        regime = determine_regime_tier(signal, weekly_state, daily_state)

        # Add QQQ-specific info
        regime["signal"] = signal
        regime["weekly_state"] = weekly_state
        regime["daily_state"] = daily_state
        regime["qqq_price"] = qqq_price
        regime["qqq_rsi"] = qqq_rsi

        # Display regime (with Windows console fallback)
        try:
            print(f"\n{regime['emoji']} MARKET REGIME: {regime['tier']}")
            print(f"   QQQ Signal: {signal}")
            print(f"   Weekly State: {weekly_state} | Daily State: {daily_state}")
            print(
                f"   QQQ Price: ${qqq_price:.2f} | RSI: {qqq_rsi:.1f}"
                if qqq_price and qqq_rsi
                else ""
            )
            print(f"\n   {regime['message']}")
        except UnicodeEncodeError:
            # Windows console fallback - skip emojis
            print(f"\nMARKET REGIME: {regime['tier']}")
            print(f"   QQQ Signal: {signal}")
            print(f"   Weekly State: {weekly_state} | Daily State: {daily_state}")
            print(
                f"   QQQ Price: ${qqq_price:.2f} | RSI: {qqq_rsi:.1f}"
                if qqq_price and qqq_rsi
                else ""
            )
            print(f"\n   {regime['message']}")
        print("=" * 80 + "\n")

        return regime

    except Exception as e:
        print(f"⚠️  ERROR analyzing QQQ: {str(e)}")
        print("⚠️  DEFAULTING TO ORANGE (CAUTIOUS) - No new buys allowed")
        return {
            "tier": "ORANGE",
            "signal": "ERROR",
            "weekly_state": "N/A",
            "daily_state": "N/A",
            "color": "orange",
            "emoji": "⚠️",
            "message": f"QQQ analysis error: {str(e)}",
            "allow_buys": False,
            "filter_mode": "exits_only",
            "qqq_price": None,
            "qqq_rsi": None,
        }


def get_score_grade(score):
    """Convert numeric score to letter grade"""
    if score >= 70:
        return "A"
    elif score >= 50:
        return "B"
    elif score >= 30:
        return "C"
    else:
        return "D"


def get_sp500_tickers():
    """Fetch current S&P 500 ticker list from Wikipedia"""
    import urllib.request

    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"

    # Add user agent to avoid 403 Forbidden
    req = urllib.request.Request(url)
    req.add_header(
        "User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    )

    try:
        with urllib.request.urlopen(req) as response:
            tables = pd.read_html(response)
            sp500_table = tables[0]
            tickers = sp500_table["Symbol"].tolist()

            # Clean up tickers (some have special characters)
            tickers = [t.replace(".", "-") for t in tickers]
            print(f"[OK] Loaded {len(tickers)} S&P 500 tickers\n")
            return tickers
    except Exception as e:
        print(f"[X] Error fetching S&P 500 list: {e}")
        print("Using fallback list of major stocks...")
        # Fallback to major stocks if Wikipedia fails
        return [
            "AAPL",
            "MSFT",
            "GOOGL",
            "AMZN",
            "NVDA",
            "META",
            "TSLA",
            "BRK-B",
            "UNH",
            "JNJ",
            "XOM",
            "V",
            "PG",
            "MA",
            "HD",
            "CVX",
            "MRK",
            "ABBV",
            "PEP",
            "COST",
        ]


def get_nasdaq100_tickers():
    """Fetch current NASDAQ 100 ticker list from Wikipedia"""
    import urllib.request

    url = "https://en.wikipedia.org/wiki/Nasdaq-100"

    # Add user agent to avoid 403 Forbidden
    req = urllib.request.Request(url)
    req.add_header(
        "User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    )

    try:
        with urllib.request.urlopen(req) as response:
            tables = pd.read_html(response)
            # Try different table indices and column names
            for table in tables:
                if "Ticker" in table.columns:
                    tickers = table["Ticker"].tolist()
                    tickers = [t.replace(".", "-") for t in tickers]
                    print(f"[OK] Loaded {len(tickers)} NASDAQ 100 tickers\n")
                    return tickers
                elif "Symbol" in table.columns:
                    tickers = table["Symbol"].tolist()
                    tickers = [t.replace(".", "-") for t in tickers]
                    print(f"[OK] Loaded {len(tickers)} NASDAQ 100 tickers\n")
                    return tickers

            # If no matching table found
            raise Exception("No table with 'Ticker' or 'Symbol' column found")
    except Exception as e:
        print(f"[X] Error fetching NASDAQ 100 list: {e}")
        print("Using fallback list of major tech stocks...")
        # Fallback to major tech stocks if Wikipedia fails
        return [
            "AAPL",
            "MSFT",
            "GOOGL",
            "GOOG",
            "AMZN",
            "NVDA",
            "META",
            "TSLA",
            "AVGO",
            "COST",
            "ASML",
            "AMD",
            "NFLX",
            "PEP",
            "ADBE",
            "CSCO",
            "TMUS",
            "CMCSA",
            "INTC",
            "TXN",
            "QCOM",
            "INTU",
            "HON",
            "AMAT",
            "AMGN",
            "BKNG",
            "ADP",
            "ISRG",
            "VRTX",
            "SBUX",
            "GILD",
            "ADI",
            "REGN",
            "LRCX",
            "MU",
            "PANW",
            "MDLZ",
            "PYPL",
            "MELI",
            "SNPS",
            "KLAC",
            "CDNS",
            "MAR",
            "MRVL",
            "CRWD",
            "ORLY",
            "CSX",
            "ADSK",
            "ABNB",
            "FTNT",
            "DASH",
            "NXPI",
            "PCAR",
            "ROP",
            "WDAY",
            "MNST",
            "CHTR",
            "CPRT",
            "PAYX",
            "AEP",
            "ROST",
            "ODFL",
            "FAST",
            "KDP",
            "EA",
            "BKR",
            "VRSK",
            "CTSH",
            "DDOG",
            "GEHC",
            "EXC",
            "XEL",
            "CCEP",
            "TEAM",
            "IDXX",
            "KHC",
            "CSGP",
            "ZS",
            "LULU",
            "ON",
            "TTWO",
            "ANSS",
            "DXCM",
            "CDW",
            "BIIB",
            "MDB",
            "WBD",
            "GFS",
            "ILMN",
            "ARM",
            "MRNA",
            "SMCI",
            "DLTR",
        ]


def get_portfolio_tickers():
    """Load portfolio tickers from stocks.txt (exclude baskets)"""
    try:
        # Go up to root, then into data folder
        stocks_file = Path(__file__).parent.parent / "data" / "stocks.txt"

        if not stocks_file.exists():
            print(f"[X] stocks.txt not found")
            return []

        tickers = []
        with open(stocks_file, "r") as f:
            for line in f:
                line = line.strip()
                # Skip comments, empty lines, and basket definitions
                if line and not line.startswith("#") and not line.startswith("["):
                    tickers.append(line)

        print(f"[OK] Loaded {len(tickers)} portfolio tickers from stocks.txt\n")
        return tickers
    except Exception as e:
        print(f"[X] Error reading stocks.txt: {e}")
        return []


def scan_stocks(
    tickers,
    category="stocks",
    daily_bars=60,
    weekly_bars=52,
    concurrency=2,
    as_of_date=None,
    regime=None,
):
    """
    Scan a list of stocks for buy opportunities

    Args:
        tickers: List of ticker symbols to scan
        category: Label for this scan (e.g., 'S&P 500', 'NASDAQ 100', 'Portfolio')
        daily_bars: Number of daily bars to analyze
        weekly_bars: Number of weekly bars to analyze
        concurrency: Number of concurrent API requests (default 2 to avoid rate limits)
        as_of_date: Optional date string (YYYY-MM-DD) for historical simulation
        regime: Market regime dict from analyze_market_regime() (optional but recommended)

    Returns:
        DataFrame with results sorted by signal strength
    """
    results = []
    buy_signals = []  # Track FULL HOLD + ADD signals
    total = len(tickers)
    completed = 0
    rate_limit_errors = 0  # Track rate limit hits

    # Display regime context if provided
    if regime:
        safe_print(
            f"{regime['emoji']} MARKET REGIME: {regime['tier']} - {regime['message']}"
        )
        print("=" * 80)

    print(f"[SCAN] Scanning {total} {category} stocks for 'FULL HOLD + ADD' signals...")
    print(
        f"Parameters: {daily_bars} daily bars, {weekly_bars} weekly bars, {concurrency} threads"
    )
    print("=" * 80)

    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        # Submit all jobs
        future_to_ticker = {
            executor.submit(
                analyze_ticker, ticker, daily_bars, weekly_bars, as_of_date
            ): ticker
            for ticker in tickers
        }

        # Process results as they complete
        for future in as_completed(future_to_ticker):
            ticker = future_to_ticker[future]
            completed += 1

            try:
                result = future.result()

                if "error" not in result:
                    results.append(result)
                    signal = result.get("signal", "N/A")

                    # Print FULL HOLD + ADD signals with entry quality
                    if signal == "FULL HOLD + ADD":
                        price = result.get("current_price", 0)
                        buy_quality = result.get("buy_quality", "N/A")
                        entry_quality = result.get("entry_quality", buy_quality)
                        entry_flag = result.get("entry_flag", "")

                        # Show entry quality (stop-aware) instead of just buy quality
                        print(
                            f"[OK] [{completed}/{total}] {ticker:6s} -> {signal:20s} ${price:,.2f} | Entry: {entry_quality:10s} {entry_flag}"
                        )
                        buy_signals.append(result)
                    # Print bearish signals with short entry quality
                    elif signal in [
                        "REDUCE",
                        "CASH",
                        "HOLD MOST + REDUCE",
                        "FULL CASH / DEFEND",
                    ]:
                        price = result.get("current_price", 0)
                        short_entry_quality = result.get("short_entry_quality", "N/A")
                        short_entry_flag = result.get("short_entry_flag", "")

                        print(
                            f"[SELL] [{completed}/{total}] {ticker:6s} -> {signal:20s} ${price:,.2f} | Entry: {short_entry_quality:10s} {short_entry_flag}"
                        )
                    elif completed % 50 == 0:  # Progress update
                        print(f"  [{completed}/{total}] Scanning...")

            except Exception as e:
                error_msg = str(e).lower()
                if (
                    "rate limit" in error_msg
                    or "429" in error_msg
                    or "too many requests" in error_msg
                ):
                    rate_limit_errors += 1
                if completed % 50 == 0:
                    print(f"  [{completed}/{total}] Scanning...")

    print("=" * 80)
    print(
        f"\n[OK] Scan complete: {len(results)} analyzed, {len(buy_signals)} FULL HOLD + ADD signals found\n"
    )

    # Rate limit warning
    if rate_limit_errors > 0:
        print(f"[!] WARNING: {rate_limit_errors} rate limit errors detected!")
        print(
            f"   Yahoo Finance may be blocking requests. Wait 5-10 minutes before next scan.\n"
        )
    elif len(results) == 0 and total > 0:
        print(f"[!] WARNING: No results returned from {total} tickers!")
        print(f"   This usually indicates Yahoo Finance rate limiting.")
        print(f"   Wait 5-10 minutes before retrying.\n")

    # Convert to DataFrame
    if not results:
        print("No results found.")
        return pd.DataFrame()

    df = pd.DataFrame(results)

    return df


def filter_buy_signals(
    df,
    signal="FULL HOLD + ADD",
    quality_filter=True,
    use_entry_quality=True,
    regime=None,
):
    """
    Filter for FULL HOLD + ADD signals with regime-aware quality filtering.

    Regime filtering modes:
    - GREEN: Normal filtering (EXCELLENT, GOOD, OK quality)
    - YELLOW: Strict filtering (EXCELLENT + SAFE ENTRY + Vol R:R >= 3.0 only)
    - ORANGE/RED: Returns empty DataFrame (no new buys allowed)

    Args:
        df: DataFrame with scan results
        signal: Signal to filter for (default: 'FULL HOLD + ADD')
        quality_filter: If True, apply quality filtering based on regime
        use_entry_quality: If True, use entry_quality (stop-aware) instead of buy_quality
        regime: Market regime dict from analyze_market_regime() (mandatory)

    Returns:
        Filtered DataFrame sorted by ticker
    """
    if df.empty:
        return df

    # MANDATORY REGIME CHECK
    if regime is None:
        raise ValueError(
            "ERROR: regime parameter is MANDATORY. Must call analyze_market_regime() first."
        )

    # ORANGE/RED: No new buys allowed
    if not regime["allow_buys"]:
        safe_print(
            f"\\n{regime['emoji']} REGIME FILTER: {regime['tier']} - No new buy signals will be shown"
        )
        print(f"   Market is in {regime['tier']} mode: {regime['message']}")
        return pd.DataFrame()  # Return empty DataFrame

    # Filter by signal
    filtered = df[df["signal"] == signal].copy()

    # Defensive validation: Ensure no bearish signals in buy results
    if not filtered.empty:
        bearish_signals = [
            "HOLD MOST + REDUCE",
            "REDUCE",
            "LIGHT / CASH",
            "CASH",
            "FULL CASH / DEFEND",
        ]
        bearish_stocks = filtered[filtered["signal"].isin(bearish_signals)]
        if not bearish_stocks.empty:
            tickers = ", ".join(bearish_stocks["ticker"].tolist())
            raise ValueError(
                f"ERROR: Bearish signal stocks found in buy signals: {tickers}. "
                "This should never happen - signal classification is broken!"
            )

    # Apply regime-based filtering
    if quality_filter and not filtered.empty:
        filter_mode = regime.get("filter_mode", "normal")

        # GREEN: Normal filtering (EXCELLENT, GOOD, OK)
        if filter_mode == "normal":
            if use_entry_quality and "entry_quality" in filtered.columns:
                filtered = filtered[
                    filtered["entry_quality"].isin(["EXCELLENT", "GOOD", "OK"])
                ]
            elif "buy_quality" in filtered.columns:
                filtered = filtered[
                    filtered["buy_quality"].isin(["EXCELLENT", "GOOD", "OK"])
                ]
            safe_print(
                f"\\n{regime['emoji']} REGIME FILTER: GREEN - Normal filtering (EXCELLENT, GOOD, OK quality)"
            )

        # YELLOW: Ultra-strict filtering (EXCELLENT + SAFE ENTRY + 3:1+ Vol R:R only)
        elif filter_mode == "strict":
            # Must have EXCELLENT entry quality
            if use_entry_quality and "entry_quality" in filtered.columns:
                filtered = filtered[filtered["entry_quality"] == "EXCELLENT"]
            elif "buy_quality" in filtered.columns:
                filtered = filtered[filtered["buy_quality"] == "EXCELLENT"]

            # Must have SAFE ENTRY flag
            if "entry_flag" in filtered.columns:
                filtered = filtered[
                    filtered["entry_flag"].str.contains(
                        "SAFE ENTRY", case=False, na=False
                    )
                ]

            # Must have Vol R:R >= 3.0
            if "vol_rr" in filtered.columns:
                filtered = filtered[filtered["vol_rr"] >= 3.0]

            safe_print(
                f"\\n{regime['emoji']} REGIME FILTER: YELLOW - Ultra-strict (EXCELLENT + SAFE ENTRY + 3:1+ Vol R:R only)"
            )
            print(f"   Market is selective: Only highest-quality setups allowed")

    return filtered.sort_values("ticker")


def filter_sell_signals(df, quality_filter=False):
    """
    Filter for bearish signals that warrant selling.
    Matches portfolio_analysis logic: ANY bearish signal = SELL

    Args:
        df: DataFrame with scan results
        quality_filter: If True, only include STRONG/MODERATE R1 quality.
                       Default False to match portfolio_analysis (all bearish signals)

    Returns:
        Filtered DataFrame sorted by ticker
    """
    if df.empty:
        return df

    # Filter for bearish signals (same as determine_portfolio_action in technical_analysis.py)
    bearish_signals = [
        "HOLD MOST + REDUCE",
        "REDUCE",
        "LIGHT / CASH",
        "CASH",
        "FULL CASH / DEFEND",
    ]
    filtered = df[df["signal"].isin(bearish_signals)].copy()

    # Defensive validation: Ensure no FULL HOLD + ADD stocks in sell signals
    if not filtered.empty:
        full_hold_stocks = filtered[filtered["signal"] == "FULL HOLD + ADD"]
        if not full_hold_stocks.empty:
            tickers = ", ".join(full_hold_stocks["ticker"].tolist())
            raise ValueError(
                f"ERROR: FULL HOLD + ADD stocks found in sell signals: {tickers}. "
                "This should never happen - signal classification is broken!"
            )

    # Optionally filter by R1 quality (disabled by default to match portfolio behavior)
    if quality_filter and "r1_quality" in filtered.columns:
        filtered = filtered[filtered["r1_quality"].isin(["STRONG", "MODERATE"])]

    return filtered.sort_values("ticker")


def cleanup_old_scans(results_dir, max_files=1, archive_retention_days=60):
    """
    Keep only the most recent scan files for EACH category (SP500, NASDAQ100, Portfolio)
    Move older ones to archive, delete archive files older than archive_retention_days

    Args:
        results_dir: Directory containing scan results
        max_files: Number of most recent files to keep per category
        archive_retention_days: Days to keep files in archive before deletion
    """
    from datetime import datetime, timedelta
    import time

    archive_dir = results_dir / "archive"
    archive_dir.mkdir(exist_ok=True)

    total_archived = 0

    # Define categories and their file patterns
    categories = [
        (
            "sp500",
            "sp500_analysis_*.xlsx",
            "scanner_report_sp500_*.pdf",
            "sp500_best_trades_*.xlsx",
            "sp500_best_trades_*.pdf",
        ),
        (
            "nasdaq100",
            "nasdaq100_analysis_*.xlsx",
            "scanner_report_nasdaq100_*.pdf",
            "nasdaq100_best_trades_*.xlsx",
            "nasdaq100_best_trades_*.pdf",
        ),
        (
            "portfolio",
            "portfolio_scanner_*.xlsx",
            "scanner_report_portfolio_*.pdf",
            "portfolio_best_trades_*.xlsx",
            "portfolio_best_trades_*.pdf",
        ),
    ]

    # Process each category separately
    for category_info in categories:
        cat_name = category_info[0]
        patterns = category_info[1:]

        # Create category subfolder and archive subfolder
        cat_dir = results_dir / cat_name
        cat_archive_dir = archive_dir / cat_name
        cat_archive_dir.mkdir(parents=True, exist_ok=True)

        # Get all files for this category
        for pattern in patterns:
            files = sorted(
                cat_dir.glob(pattern),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )

            # Move files beyond max_files to archive
            for old_file in files[max_files:]:
                archive_path = cat_archive_dir / old_file.name
                old_file.rename(archive_path)
                print(f"  [ARCHIVED] ({cat_name}): {old_file.name}")
                total_archived += 1

    if total_archived > 0:
        print(f"  [OK] Archived {total_archived} file(s), kept {max_files} most recent")
    else:
        print(f"  [OK] No files to archive (only {max_files} most recent exist)")

    # Delete archive files older than retention period
    cutoff_time = time.time() - (
        archive_retention_days * 86400
    )  # 86400 seconds per day
    deleted_count = 0

    # Check all category subfolders in archive
    for cat_archive_dir in archive_dir.glob("*/"):
        if cat_archive_dir.is_dir():
            for archive_file in cat_archive_dir.glob("*"):
                if (
                    archive_file.is_file()
                    and archive_file.stat().st_mtime < cutoff_time
                ):
                    try:
                        archive_file.unlink()
                        deleted_count += 1
                    except Exception as e:
                        print(f"  [WARNING] Could not delete {archive_file.name}: {e}")

    if deleted_count > 0:
        print(
            f"  [DELETED] {deleted_count} archive file(s) older than {archive_retention_days} days"
        )


def calculate_buy_score(row):
    """
    Calculate trading score for buy opportunities (0-100)

    Scoring factors:
    - Entry Quality (25pts): EXCELLENT=25, GOOD=18, OK=12 (uses stop-aware quality)
    - Distance to S1 (15pts): <2%=15, 2-5%=11, 5-10%=7
    - Upside to R1 (20pts): >15%=20, 10-15%=16, 5-10%=12, 2-5%=8
    - Risk/Reward (20pts): >3:1=20, 2-3:1=15, 1-2:1=10, 0.5-1:1=5, <0.5:1=0
    - RSI (12pts): <30=12, 30-40=9, 40-50=6
    - BB Position (8pts): Below lower=8, touching=6, near=4
    """
    score = 0

    # Entry Quality score (25 points) - use stop-aware quality
    # Prefer entry_quality (stop-aware) over buy_quality
    quality = row.get("entry_quality", row.get("buy_quality", "")).upper()
    if quality == "EXCELLENT":
        score += 25
    elif quality == "GOOD":
        score += 18
    elif quality == "OK":
        score += 12

    # Distance to S1 (15 points)
    price = row.get("current_price", 0)
    s1 = row.get("s1", 0)
    if price > 0 and s1 > 0:
        distance_pct = abs((price - s1) / price * 100)
        if distance_pct < 2:
            score += 15
        elif distance_pct < 5:
            score += 11
        elif distance_pct < 10:
            score += 7

    # Upside potential to R1 (20 points)
    r1 = row.get("r1", 0)
    upside_pct = 0
    if price > 0 and r1 > 0 and r1 > price:
        upside_pct = ((r1 - price) / price) * 100
        if upside_pct > 15:
            score += 20
        elif upside_pct > 10:
            score += 16
        elif upside_pct > 5:
            score += 12
        elif upside_pct > 2:
            score += 8

    # Risk/Reward Ratio (20 points) - NEW!
    risk_pct = abs((price - s1) / price * 100) if price > 0 and s1 > 0 else 0
    # Apply minimum risk threshold to prevent unrealistic R/R when support is too close
    risk_pct = max(risk_pct, 2.0)  # Minimum 2% risk
    if upside_pct > 0 and risk_pct > 0:
        rr_ratio = upside_pct / risk_pct  # Reward:Risk
        if rr_ratio >= 3:
            score += 20
        elif rr_ratio >= 2:
            score += 15
        elif rr_ratio >= 1:
            score += 10
        elif rr_ratio >= 0.5:
            score += 5

    # RSI score (12 points)
    rsi = row.get("rsi")
    if rsi is not None:
        if rsi < 30:
            score += 12
        elif rsi < 40:
            score += 9
        elif rsi < 50:
            score += 6

    # Bollinger Band position (8 points)
    bb_lower = row.get("bb_lower")
    if price > 0 and bb_lower is not None and bb_lower > 0:
        bb_pct = (price - bb_lower) / bb_lower * 100
        if bb_pct < 0:  # Below lower band
            score += 8
        elif bb_pct < 1:  # Touching lower band
            score += 6
        elif bb_pct < 3:  # Near lower band
            score += 4

    return score


def calculate_sell_score(row):
    """
    Calculate trading score for sell opportunities (0-100)

    Scoring factors:
    - Quality (25pts): EXCELLENT=25, GOOD=18, OK=12
    - Distance to R1 (15pts): <2%=15, 2-5%=11, 5-10%=7
    - Profit to S1 (20pts): >15%=20, 10-15%=16, 5-10%=12, 2-5%=8
    - Risk/Reward (20pts): >3:1=20, 2-3:1=15, 1-2:1=10, 0.5-1:1=5, <0.5:1=0
    - RSI (12pts): >70=12, 60-70=9, 50-60=6
    - BB Position (8pts): Above upper=8, touching=6, near=4
    """
    score = 0

    # Quality score (25 points)
    quality = row.get("r1_quality", "").upper()
    if quality == "EXCELLENT":
        score += 25
    elif quality == "GOOD":
        score += 18
    elif quality == "OK":
        score += 12

    # Distance to R1 (15 points)
    price = row.get("current_price", 0)
    r1 = row.get("r1", 0)
    if price > 0 and r1 > 0:
        distance_pct = abs((r1 - price) / price * 100)
        if distance_pct < 2:
            score += 15
        elif distance_pct < 5:
            score += 11
        elif distance_pct < 10:
            score += 7

    # Profit potential to S1 (20 points)
    s1 = row.get("s1", 0)
    profit_pct = 0
    if price > 0 and s1 > 0 and price > s1:
        profit_pct = ((price - s1) / price) * 100
        if profit_pct > 15:
            score += 20
        elif profit_pct > 10:
            score += 16
        elif profit_pct > 5:
            score += 12
        elif profit_pct > 2:
            score += 8

    # Risk/Reward Ratio (20 points) - NEW!
    risk_pct = (
        abs((r1 - price) / price * 100) if price > 0 and r1 > 0 and r1 > price else 0
    )
    # Apply minimum risk threshold to prevent unrealistic R/R when resistance is too close
    risk_pct = max(risk_pct, 2.0)  # Minimum 2% risk
    if profit_pct > 0 and risk_pct > 0:
        rr_ratio = profit_pct / risk_pct  # Reward:Risk
        if rr_ratio >= 3:
            score += 20
        elif rr_ratio >= 2:
            score += 15
        elif rr_ratio >= 1:
            score += 10
        elif rr_ratio >= 0.5:
            score += 5

    # RSI score (12 points)
    rsi = row.get("rsi")
    if rsi is not None:
        if rsi > 70:
            score += 12
        elif rsi > 60:
            score += 9
        elif rsi > 50:
            score += 6

    # Bollinger Band position (8 points)
    bb_upper = row.get("bb_upper")
    if price > 0 and bb_upper is not None and bb_upper > 0:
        bb_pct = (price - bb_upper) / bb_upper * 100
        if bb_pct > 0:  # Above upper band
            score += 8
        elif bb_pct > -1:  # Touching upper band
            score += 6
        elif bb_pct > -3:  # Near upper band
            score += 5

    return score


def create_best_trades_excel(buy_df, sell_df, output_file, category=""):
    """
    Create Excel with ranked best trading opportunities

    Args:
        buy_df: DataFrame with buy opportunities
        sell_df: DataFrame with sell opportunities
        output_file: Path to save Excel file
        category: Label (e.g., 'S&P 500', 'NASDAQ 100')
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styling
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    # === TAB 1: TOP BUY SETUPS ===
    if not buy_df.empty:
        # OPTION 1: Filter for EXCELLENT and GOOD quality only, then rank by Vol R:R
        buy_df = buy_df.copy()

        # Filter for quality setups only - use entry_quality (stop-aware)
        quality_col = (
            "entry_quality" if "entry_quality" in buy_df.columns else "buy_quality"
        )
        buy_df = buy_df[buy_df[quality_col].isin(["EXCELLENT", "GOOD"])].copy()

        if not buy_df.empty:
            # Calculate Vol R:R for ranking
            def calc_vol_rr(row):
                price = row.get("current_price", 0)
                r1 = row.get("r1", 0)
                suggested_stop = row.get("suggested_stop_pct", 5.0)

                reward = (
                    ((r1 - price) / price * 100)
                    if price > 0 and r1 > 0 and r1 > price
                    else 0
                )
                return (
                    reward / suggested_stop if suggested_stop > 0 and reward > 0 else 0
                )

            buy_df["vol_rr"] = buy_df.apply(calc_vol_rr, axis=1)

            # Sort by Vol R:R (highest first), take top 20
            buy_df = buy_df.sort_values("vol_rr", ascending=False).head(20)
            buy_df["rank"] = range(1, len(buy_df) + 1)

            ws_buy = wb.create_sheet("Top Buy Setups")

            # Title
            ws_buy.cell(1, 1, f"{category} - BEST BUY SETUPS").font = Font(
                bold=True, size=14
            )
            ws_buy.cell(
                2,
                1,
                f"Top {len(buy_df)} EXCELLENT/GOOD entry quality (stop-aware) ranked by Vol R:R",
            ).font = Font(italic=True, color="666666")

            # Headers - Simplified for volatility-based trading
            headers = [
                "Rank",
                "Ticker",
                "Signal",
                "Quality",
                "Quality Flag",
                "Current Price",
                "Vol R:R",
                "Vol Stop Price",
                "Vol Stop Loss %",
                "Target R1",
                "Target R1 Gain %",
                "Support S1",
                "Distance to S1",
                "RSI",
                "BB %",
                "Daily Range %",
                "Volatility",
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws_buy.cell(4, col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            # Data rows
            row = 5
            for _, stock in buy_df.iterrows():
                price = stock.get("current_price", 0)
                s1 = stock.get("s1", 0)
                r1 = stock.get("r1", 0)
                bb_lower = stock.get("bb_lower", 0)

                # Calculate key metrics
                reward_pct = (
                    ((r1 - price) / price * 100)
                    if price > 0 and r1 > 0 and r1 > price
                    else None
                )
                bb_pct = (price - bb_lower) / bb_lower * 100 if bb_lower > 0 else None
                distance_to_s1 = (
                    abs((price - s1) / price * 100) if price > 0 and s1 > 0 else None
                )

                # Volatility stop calculations
                suggested_stop = stock.get("suggested_stop_pct", 5.0)
                vol_stop_price = price * (1 - suggested_stop / 100)
                vol_rr = stock.get("vol_rr", 0)

                # Write simplified row data
                ws_buy.cell(row, 1, stock["rank"])
                ws_buy.cell(row, 2, stock["ticker"])
                ws_buy.cell(row, 3, stock.get("signal", ""))
                # Use entry_quality (stop-aware) if available, otherwise buy_quality
                quality_display = stock.get(
                    "entry_quality", stock.get("buy_quality", "")
                )
                entry_flag = stock.get("entry_flag", "")
                ws_buy.cell(row, 4, quality_display)
                ws_buy.cell(row, 5, entry_flag)
                ws_buy.cell(row, 6, price)
                ws_buy.cell(row, 7, f"1:{vol_rr:.1f}" if vol_rr > 0 else "N/A")
                ws_buy.cell(row, 8, vol_stop_price)
                ws_buy.cell(row, 9, f"-{suggested_stop:.1f}%")
                ws_buy.cell(row, 10, r1 if r1 > 0 else "")
                ws_buy.cell(row, 11, f"+{reward_pct:.1f}%" if reward_pct else "")
                ws_buy.cell(row, 12, s1)
                ws_buy.cell(row, 13, f"{distance_to_s1:.1f}%" if distance_to_s1 else "")
                ws_buy.cell(
                    row,
                    14,
                    f"{stock.get('rsi'):.1f}" if stock.get("rsi") is not None else "",
                )
                ws_buy.cell(row, 15, f"{bb_pct:.1f}%" if bb_pct else "")
                daily_range = stock.get("avg_daily_range_pct")
                ws_buy.cell(row, 16, f"{daily_range:.2f}%" if daily_range else "")
                ws_buy.cell(row, 17, stock.get("volatility_class", ""))
                row += 1

            # Auto-size columns
            for col in ws_buy.columns:
                max_length = 12
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_buy.column_dimensions[col[0].column_letter].width = min(
                    max_length + 2, 50
                )

    # === TAB 2: TOP SELL SETUPS ===
    if not sell_df.empty:
        # OPTION 1: Filter for EXCELLENT and GOOD exit quality (short entry), then rank by Vol R:R
        sell_df = sell_df.copy()

        # Filter for quality setups only (use exit_quality if available, fallback to r1_quality)
        quality_col = (
            "exit_quality" if "exit_quality" in sell_df.columns else "r1_quality"
        )
        sell_df = sell_df[sell_df[quality_col].isin(["EXCELLENT", "GOOD"])].copy()

        if not sell_df.empty:
            # Calculate Vol R:R for ranking
            def calc_vol_rr(row):
                price = row.get("current_price", 0)
                s1 = row.get("s1", 0)
                suggested_stop = row.get("suggested_stop_pct", 5.0)

                reward = (
                    ((price - s1) / price * 100)
                    if price > 0 and s1 > 0 and price > s1
                    else 0
                )
                return (
                    reward / suggested_stop if suggested_stop > 0 and reward > 0 else 0
                )

            sell_df["vol_rr"] = sell_df.apply(calc_vol_rr, axis=1)

            # Sort by Vol R:R (highest first), take top 20
            sell_df = sell_df.sort_values("vol_rr", ascending=False).head(20)
            sell_df["rank"] = range(1, len(sell_df) + 1)

            ws_sell = wb.create_sheet("Top Sell Setups")

            # Title
            ws_sell.cell(1, 1, f"{category} - BEST SELL SETUPS").font = Font(
                bold=True, size=14
            )
            ws_sell.cell(
                2, 1, f"Top {len(sell_df)} EXCELLENT/GOOD quality ranked by Vol R:R"
            ).font = Font(italic=True, color="666666")

            # Headers - Simplified for volatility-based trading
            headers = [
                "Rank",
                "Ticker",
                "Signal",
                "Quality",
                "Quality Flag",
                "Current Price",
                "Vol R:R",
                "Vol Stop Price",
                "Vol Stop Loss %",
                "Target S1",
                "Target S1 Gain %",
                "Resistance R1",
                "Distance to R1",
                "RSI",
                "BB %",
                "Daily Range %",
                "Volatility",
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws_sell.cell(4, col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            # Data rows
            row = 5
            for _, stock in sell_df.iterrows():
                price = stock.get("current_price", 0)
                r1 = stock.get("r1", 0)
                s1 = stock.get("s1", 0)
                bb_upper = stock.get("bb_upper", 0)

                # Calculate key metrics (for sells: profit down to S1, risk up to vol stop)
                reward_pct = (
                    ((price - s1) / price * 100)
                    if price > 0 and s1 > 0 and price > s1
                    else None
                )
                bb_pct = (price - bb_upper) / bb_upper * 100 if bb_upper > 0 else None
                distance_to_r1 = (
                    abs((r1 - price) / price * 100) if price > 0 and r1 > 0 else None
                )

                # Volatility stop calculations
                suggested_stop = stock.get("suggested_stop_pct", 5.0)
                vol_stop_price = price * (1 + suggested_stop / 100)
                vol_rr = stock.get("vol_rr", 0)

                # Write simplified row data
                ws_sell.cell(row, 1, stock["rank"])
                ws_sell.cell(row, 2, stock["ticker"])
                ws_sell.cell(row, 3, stock.get("signal", ""))

                # Use exit_quality and short_entry_flag for short entry assessment
                quality_display = stock.get("exit_quality", stock.get("r1_quality", ""))
                short_entry_flag = stock.get("short_entry_flag", "")
                ws_sell.cell(row, 4, quality_display)
                ws_sell.cell(row, 5, short_entry_flag)
                ws_sell.cell(row, 6, price)
                ws_sell.cell(row, 7, f"1:{vol_rr:.1f}" if vol_rr > 0 else "N/A")
                ws_sell.cell(row, 8, vol_stop_price)
                ws_sell.cell(row, 9, f"-{suggested_stop:.1f}%")
                ws_sell.cell(row, 10, s1 if s1 > 0 else "")
                ws_sell.cell(row, 11, f"+{reward_pct:.1f}%" if reward_pct else "")
                ws_sell.cell(row, 12, r1)
                ws_sell.cell(
                    row, 13, f"{distance_to_r1:.1f}%" if distance_to_r1 else ""
                )
                ws_sell.cell(
                    row,
                    14,
                    f"{stock.get('rsi'):.1f}" if stock.get("rsi") is not None else "",
                )
                ws_sell.cell(row, 15, f"{bb_pct:.1f}%" if bb_pct else "")
                daily_range = stock.get("avg_daily_range_pct")
                ws_sell.cell(row, 16, f"{daily_range:.2f}%" if daily_range else "")
                ws_sell.cell(row, 17, stock.get("volatility_class", ""))
                row += 1

            # Auto-size columns
            for col in ws_sell.columns:
                max_length = 12
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_sell.column_dimensions[col[0].column_letter].width = min(
                    max_length + 2, 50
                )

    # === TAB 3: SUMMARY ===
    ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning
    ws_summary.cell(1, 1, f"{category} - BEST TRADES SUMMARY").font = Font(
        bold=True, size=14
    )
    ws_summary.cell(
        3, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ).font = Font(italic=True)

    ws_summary.cell(5, 1, "Buy Setups:").font = Font(bold=True)
    ws_summary.cell(5, 2, len(buy_df) if not buy_df.empty else 0)

    ws_summary.cell(6, 1, "Sell Setups:").font = Font(bold=True)
    ws_summary.cell(6, 2, len(sell_df) if not sell_df.empty else 0)

    if not buy_df.empty:
        ws_summary.cell(8, 1, "Top Buy Vol R:R:").font = Font(bold=True)
        ws_summary.cell(8, 2, f"{buy_df['vol_rr'].max():.2f}:1")
        ws_summary.cell(9, 1, "Avg Buy Vol R:R:").font = Font(bold=True)
        ws_summary.cell(9, 2, f"{buy_df['vol_rr'].mean():.2f}:1")

    if not sell_df.empty:
        ws_summary.cell(11, 1, "Top Sell Vol R:R:").font = Font(bold=True)
        ws_summary.cell(11, 2, f"{sell_df['vol_rr'].max():.2f}:1")
        ws_summary.cell(12, 1, "Avg Sell Vol R:R:").font = Font(bold=True)
        ws_summary.cell(12, 2, f"{sell_df['vol_rr'].mean():.2f}:1")

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20

    # Save workbook
    wb.save(output_file)
    print(f"[OK] Best Trades Excel created: {output_file}")


def create_excel_output(buy_df, sell_df, output_file, category=""):
    """
    Create Excel workbook with tabs for buy and sell opportunities.
    Matches portfolio_tracker format from portfolio_reports.py

    Args:
        buy_df: DataFrame with FULL HOLD + ADD results (quality-filtered)
        sell_df: DataFrame with bearish signal results
        output_file: Path to output Excel file
        category: Category label (e.g., 'S&P 500', 'NASDAQ 100')
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Define styles (matching portfolio_reports)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    # === TAB 1: BUY SETUPS ===
    if not buy_df.empty:
        ws_buy = wb.create_sheet("Buy Setups")

        # Title
        ws_buy.cell(1, 1, f"{category} - BUY SETUPS").font = Font(bold=True, size=14)
        ws_buy.cell(
            2, 1, f"FULL HOLD + ADD signals with EXCELLENT/GOOD/OK quality"
        ).font = Font(italic=True, color="666666")

        # Headers at row 4
        buy_headers = [
            "Ticker",
            "Price",
            "Signal",
            "Quality",
            "Quality Flag",
            "RSI",
            "BB_Upper",
            "BB_Middle",
            "BB_Lower",
            "S1",
            "S2",
            "S3",
            "R1",
            "R2",
            "R3",
            "D50 MA",
            "D100 MA",
            "D200 MA",
            "Stop Level (8%)",
            "Accessible Supports",
        ]

        for col_num, header in enumerate(buy_headers, 1):
            cell = ws_buy.cell(4, col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Data rows
        row = 5
        for _, stock in buy_df.sort_values("ticker").iterrows():
            ws_buy.cell(row, 1, stock["ticker"])
            ws_buy.cell(row, 2, stock.get("current_price"))
            ws_buy.cell(row, 3, stock["signal"])
            ws_buy.cell(row, 4, stock.get("entry_quality", stock.get("buy_quality")))
            ws_buy.cell(row, 5, stock.get("entry_flag", ""))
            ws_buy.cell(row, 6, stock.get("rsi"))
            ws_buy.cell(row, 7, stock.get("bb_upper"))
            ws_buy.cell(row, 8, stock.get("bb_middle"))
            ws_buy.cell(row, 9, stock.get("bb_lower"))
            ws_buy.cell(row, 10, stock.get("s1"))
            ws_buy.cell(row, 11, stock.get("s2"))
            ws_buy.cell(row, 12, stock.get("s3"))
            ws_buy.cell(row, 13, stock.get("r1"))
            ws_buy.cell(row, 14, stock.get("r2"))
            ws_buy.cell(row, 15, stock.get("r3"))
            ws_buy.cell(row, 16, stock.get("d50"))
            ws_buy.cell(row, 17, stock.get("d100"))
            ws_buy.cell(row, 18, stock.get("d200"))
            ws_buy.cell(row, 19, stock.get("stop_level"))
            ws_buy.cell(row, 20, stock.get("accessible_supports_count"))
            row += 1

        # Auto-size columns
        for col in ws_buy.columns:
            max_length = 12
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_buy.column_dimensions[col[0].column_letter].width = min(
                max_length + 2, 50
            )

    # === TAB 2: SELL SETUPS ===
    if not sell_df.empty:
        ws_sell = wb.create_sheet("Sell Setups")

        # Title
        ws_sell.cell(1, 1, f"{category} - SELL SETUPS").font = Font(bold=True, size=14)
        ws_sell.cell(2, 1, f"Bearish signals - reduce exposure").font = Font(
            italic=True, color="666666"
        )

        # Headers at row 4
        sell_headers = [
            "Ticker",
            "Price",
            "Signal",
            "Quality",
            "Quality Flag",
            "RSI",
            "BB_Upper",
            "BB_Middle",
            "BB_Lower",
            "R1 Quality",
            "R1",
            "R2",
            "R3",
            "D50 MA",
            "D100 MA",
            "D200 MA",
        ]

        for col_num, header in enumerate(sell_headers, 1):
            cell = ws_sell.cell(4, col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Data rows
        row = 5
        for _, stock in sell_df.sort_values("ticker").iterrows():
            ws_sell.cell(row, 1, stock["ticker"])
            ws_sell.cell(row, 2, stock.get("current_price"))
            ws_sell.cell(row, 3, stock["signal"])
            ws_sell.cell(row, 4, stock.get("short_entry_quality", "N/A"))
            ws_sell.cell(row, 5, stock.get("short_entry_flag", ""))
            ws_sell.cell(row, 6, stock.get("rsi"))
            ws_sell.cell(row, 7, stock.get("bb_upper"))
            ws_sell.cell(row, 8, stock.get("bb_middle"))
            ws_sell.cell(row, 9, stock.get("bb_lower"))
            ws_sell.cell(row, 10, stock.get("r1_quality"))
            ws_sell.cell(row, 11, stock.get("r1"))
            ws_sell.cell(row, 12, stock.get("r2"))
            ws_sell.cell(row, 13, stock.get("r3"))
            ws_sell.cell(row, 14, stock.get("d50"))
            ws_sell.cell(row, 15, stock.get("d100"))
            ws_sell.cell(row, 16, stock.get("d200"))
            row += 1

        # Auto-size columns
        for col in ws_sell.columns:
            max_length = 12
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_sell.column_dimensions[col[0].column_letter].width = min(
                max_length + 2, 50
            )

    # Save workbook
    wb.save(output_file)

    print(f"[OK] Excel file created: {output_file}")
    if not buy_df.empty:
        print(f"  - Buy Setups: {len(buy_df)} stocks")
    if not sell_df.empty:
        print(f"  - Sell Setups: {len(sell_df)} stocks")


def create_portfolio_excel(all_df, output_file, category="Portfolio"):
    """
    Create Excel workbook for portfolio with ALL stocks organized by signal

    Args:
        all_df: DataFrame with ALL portfolio results (not filtered)
        output_file: Path to output Excel file
        category: Category label (default: 'Portfolio')
    """
    if all_df.empty:
        print(f"No data to export to Excel for {category}")
        return

    # Define column order (only columns used in portfolio PDF report)
    columns = [
        "ticker",
        "signal",
        "current_price",
        "s1",
        "s2",
        "s3",  # Support levels (Zones)
        "r1",
        "r2",
        "r3",  # Resistance levels (Targets)
    ]

    # Keep only columns that exist in the dataframe
    available_cols = [col for col in columns if col in all_df.columns]

    # Sort alphabetically by ticker
    all_df = all_df.sort_values("ticker")

    # Filter by signal (matching portfolio PDF format exactly)
    df_full_hold_add = all_df[all_df["signal"] == "FULL HOLD + ADD"][
        available_cols
    ].copy()
    df_hold_reduce = all_df[all_df["signal"] == "HOLD MOST + REDUCE"][
        available_cols
    ].copy()
    df_hold = all_df[all_df["signal"] == "HOLD"][available_cols].copy()
    df_cash = all_df[all_df["signal"] == "CASH"][available_cols].copy()
    df_all = all_df[available_cols].copy()

    # Create Excel writer
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Sheet 1: All stocks
        df_all.to_excel(writer, sheet_name="All", index=False)

        # Sheet 2: FULL HOLD + ADD (detailed in PDF)
        df_full_hold_add.to_excel(writer, sheet_name="FULL HOLD + ADD", index=False)

        # Sheet 3: HOLD MOST + REDUCE (summary in PDF)
        df_hold_reduce.to_excel(writer, sheet_name="HOLD + REDUCE", index=False)

        # Sheet 4: HOLD (summary in PDF)
        df_hold.to_excel(writer, sheet_name="HOLD", index=False)

        # Sheet 5: CASH (summary in PDF)
        df_cash.to_excel(writer, sheet_name="CASH", index=False)

        # Auto-adjust column widths for all sheets
        for sheet_name in writer.sheets:
            if sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (AttributeError, TypeError) as e:
                            logger.debug(
                                f"Skipping cell value for width calculation: {e}"
                            )
                        except Exception as e:
                            logger.warning(
                                f"Unexpected error calculating column width: {e}"
                            )
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"âœ“ Excel file created: {output_file}")
    print(f"  - All: {len(df_all)} stocks")
    print(f"  - FULL HOLD + ADD: {len(df_full_hold_add)} stocks")
    print(f"  - HOLD MOST + REDUCE: {len(df_hold_reduce)} stocks")
    print(f"  - HOLD: {len(df_hold)} stocks")
    print(f"  - CASH: {len(df_cash)} stocks")


def create_pdf_report(
    buy_df, sell_df, output_file, timestamp_str, category="", regime=None
):
    """
    Create comprehensive PDF trading report matching portfolio_reports format.
    Adapted for scanner (no holdings/targets data)

    Args:
        buy_df: DataFrame with quality-filtered buy signals
        sell_df: DataFrame with sell signals
        output_file: Path to output PDF file
        timestamp_str: Timestamp string for report header
        category: Category label (e.g., 'S&P 500', 'NASDAQ 100')
        regime: Market regime dict from analyze_market_regime() (optional, for display)
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        PageBreak,
    )
    from reportlab.lib.enums import TA_CENTER

    if buy_df.empty and sell_df.empty:
        print(f"No data to create PDF report for {category}")
        return

    doc = SimpleDocTemplate(
        str(output_file),
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles (matching portfolio_reports)
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#2c3e50"),
        spaceAfter=20,
        alignment=TA_CENTER,
    )

    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=16,
        textColor=colors.HexColor("#2980b9"),
        spaceAfter=12,
        spaceBefore=20,
    )

    subsection_style = ParagraphStyle(
        "Subsection",
        parent=styles["Heading3"],
        fontSize=12,
        textColor=colors.HexColor("#34495e"),
        spaceAfter=6,
        spaceBefore=10,
    )

    # === PAGE 1: DASHBOARD ===
    report_title = category.replace("&", "&amp;")
    cover_title = f"Trading Playbook - {report_title}<br/><font size=14>{datetime.now().strftime('%B %d, %Y')}</font>"
    elements.append(Paragraph(cover_title, title_style))
    elements.append(Spacer(1, 0.3 * inch))

    # Market Regime Banner (if provided)
    if regime:
        regime_color = {
            "GREEN": "#27ae60",  # green
            "YELLOW": "#f39c12",  # yellow/orange
            "ORANGE": "#e67e22",  # orange
            "RED": "#c0392b",  # red
        }.get(
            regime["tier"], "#95a5a6"
        )  # gray default

        regime_banner = f"""
        <para align=center bgcolor='{regime_color}' textColor='white'>
        <b><font size=14>{regime['emoji']} MARKET REGIME: {regime['tier']}</font></b><br/>
        <font size=11>QQQ Signal: {regime['signal']} | Weekly: {regime['weekly_state']} | Daily: {regime['daily_state']}</font><br/>
        <font size=10>{regime['message']}</font>
        </para>
        """
        elements.append(Paragraph(regime_banner, styles["Normal"]))
        elements.append(Spacer(1, 0.2 * inch))

    # Scanner summary
    summary_text = f"""
    <b>Market Scan Summary</b><br/>
    <br/>
    <b>Buy Setups:</b> {len(buy_df)} stocks (FULL HOLD + ADD with quality)<br/>
    <b>Sell Setups:</b> {len(sell_df)} stocks (Bearish - reduce exposure)<br/>
    <br/>
    <b>Report Contents:</b><br/>
    - Section 1: Buy Setups (detailed entry plans)<br/>
    - Section 2: Sell Setups (exit recommendations)<br/>
    <br/>
    <i>Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}</i>
    """
    elements.append(Paragraph(summary_text, styles["Normal"]))
    elements.append(PageBreak())

    # === SECTION 1: BUY SETUPS ===
    if not buy_df.empty:
        elements.append(Paragraph(f"BUY SETUPS ({len(buy_df)} stocks)", section_style))
        elements.append(Spacer(1, 0.1 * inch))

        # Sort by vol_rr if available, otherwise by ticker
        sort_col = "vol_rr" if "vol_rr" in buy_df.columns else "ticker"
        sort_asc = False if sort_col == "vol_rr" else True

        for rank, (_, stock) in enumerate(
            buy_df.sort_values(sort_col, ascending=sort_asc).iterrows(), 1
        ):
            ticker = stock["ticker"]
            price = stock["current_price"]
            signal = stock["signal"]
            buy_quality = stock.get("buy_quality", "N/A")
            rsi = stock.get("rsi")

            # Get stop-aware entry quality
            entry_quality = stock.get("entry_quality", buy_quality)
            entry_flag = stock.get("entry_flag", "")
            entry_note = stock.get("entry_note", "")
            stop_tolerance = stock.get("stop_tolerance_pct", 8.0)
            stop_level = stock.get("stop_level", 0)
            vol_rr = stock.get("vol_rr", 0)

            # Color code the R:R ratio
            if vol_rr >= 3:
                rr_color = "#27ae60"  # Green
            elif vol_rr >= 2:
                rr_color = "#3498db"  # Blue
            elif vol_rr >= 1:
                rr_color = "#f39c12"  # Orange
            else:
                rr_color = "#e74c3c"  # Red

            # Stock header with ranking, entry quality and flag
            entry_flag_text = f" {entry_flag}" if entry_flag else ""
            header_text = f"<b>#{rank}. {ticker} - ${price:,.2f}</b> | <font color='#2980b9'><b>{entry_quality}{entry_flag_text}</b></font> | Vol R:R: <font color='{rr_color}'><b>{vol_rr:.1f}:1</b></font>"
            elements.append(Paragraph(header_text, subsection_style))

            # RSI and Signal line
            rsi_text = f"{rsi:.1f}" if rsi else "N/A"
            signal_line = f"<b>RSI:</b> {rsi_text} | <b>Signal:</b> {signal}"
            elements.append(Paragraph(signal_line, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Volatility Stop and Support Info
            suggested_stop = stock.get("suggested_stop_pct", 5.0)
            vol_stop_price = price * (1 - suggested_stop / 100)
            accessible_supports = stock.get("accessible_supports_count", 0)

            stop_info = f"""
            <b>Volatility Stop ({suggested_stop:.1f}%):</b> ${vol_stop_price:.2f} loss = <font color="#e74c3c">-{suggested_stop:.1f}%</font><br/>
            <br/>
            <b>{stop_tolerance:.0f}% Stop Tolerance:</b> ${stop_level:.2f} | <b>Accessible Supports:</b> {accessible_supports} within range<br/>
            """
            elements.append(Paragraph(stop_info, styles["Normal"]))

            # Target R1
            r1 = stock.get("r1", 0)
            if r1 and price > 0:
                reward_pct = ((r1 - price) / price * 100) if r1 > price else 0
                target_text = f"<b>Target R1:</b> ${r1:.2f} gain = <font color='#27ae60'>+{reward_pct:.1f}%</font>"
                elements.append(Paragraph(target_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Support S1 and Volatility info
            s1 = stock.get("s1", 0)
            volatility_class = stock.get("volatility_class", "")
            daily_range = stock.get("avg_daily_range_pct", 0)
            if s1 and price > 0:
                distance_to_s1 = abs((price - s1) / price * 100)
                s1_text = f"<b>Support S1:</b> ${s1:.2f} ({distance_to_s1:.1f}% away) | <b>Volatility:</b> {volatility_class} (~{daily_range:.2f}% daily range)"
                elements.append(Paragraph(s1_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Entry note
            if entry_note:
                elements.append(Paragraph(f"<i>{entry_note}</i>", styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Support levels (entry zones)
            s1 = stock.get("s1")
            s2 = stock.get("s2")
            s3 = stock.get("s3")

            support_text = "<b>Entry Zones (Support Levels):</b><br/>"
            if s1:
                support_text += f"- S1: ${s1:,.2f}<br/>"
            if s2:
                support_text += f"- S2: ${s2:,.2f}<br/>"
            if s3:
                support_text += f"- S3: ${s3:,.2f}<br/>"
            elements.append(Paragraph(support_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Buy quality for each support level (matching portfolio format)
            s1_quality = stock.get("s1_quality", "N/A")
            s1_quality_note = stock.get("s1_quality_note", "")
            s2_quality = stock.get("s2_quality", "N/A")
            s2_quality_note = stock.get("s2_quality_note", "")
            s3_quality = stock.get("s3_quality", "N/A")
            s3_quality_note = stock.get("s3_quality_note", "")

            quality_text = f"""
            <b>Buy Quality S1:</b> {s1_quality} - {s1_quality_note}<br/>
            <b>Buy Quality S2:</b> {s2_quality} - {s2_quality_note}<br/>
            <b>Buy Quality S3:</b> {s3_quality} - {s3_quality_note}
            """
            elements.append(Paragraph(quality_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Resistance levels (targets)
            r1 = stock.get("r1")
            r2 = stock.get("r2")
            r3 = stock.get("r3")

            target_text = "<b>Upside Targets (Resistance Levels):</b><br/>"
            if r1:
                gain1 = (r1 / price - 1) * 100 if price > 0 else 0
                target_text += f"- R1: ${r1:,.2f} ({gain1:.1f}% gain)<br/>"
            if r2:
                gain2 = (r2 / price - 1) * 100 if price > 0 else 0
                target_text += f"- R2: ${r2:,.2f} ({gain2:.1f}% gain)<br/>"
            if r3:
                gain3 = (r3 / price - 1) * 100 if price > 0 else 0
                target_text += f"- R3: ${r3:,.2f} ({gain3:.1f}% gain)<br/>"
            elements.append(Paragraph(target_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Sell quality for resistance levels (exit quality assessment)
            r1_quality = stock.get("r1_quality", "N/A")
            r1_quality_note = stock.get("r1_quality_note", "")
            r2_quality = stock.get("r2_quality", "N/A")
            r2_quality_note = stock.get("r2_quality_note", "")
            r3_quality = stock.get("r3_quality", "N/A")
            r3_quality_note = stock.get("r3_quality_note", "")

            sell_quality_text = f"""
            <b>Sell Quality R1:</b> {r1_quality} - {r1_quality_note}<br/>
            <b>Sell Quality R2:</b> {r2_quality} - {r2_quality_note}<br/>
            <b>Sell Quality R3:</b> {r3_quality} - {r3_quality_note}
            """
            elements.append(Paragraph(sell_quality_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Moving averages
            d50 = stock.get("d50")
            d100 = stock.get("d100")
            d200 = stock.get("d200")
            if d50 or d100 or d200:
                ma_text = "<b>Moving Averages:</b> "
                ma_parts = []
                if d50:
                    ma_parts.append(f"D50 ${d50:,.2f}")
                if d100:
                    ma_parts.append(f"D100 ${d100:,.2f}")
                if d200:
                    ma_parts.append(f"D200 ${d200:,.2f}")
                ma_text += " | ".join(ma_parts)
                elements.append(Paragraph(ma_text, styles["Normal"]))

            # Volume Profile (matching portfolio format)
            poc = stock.get("poc_60d")
            val = stock.get("val_60d")
            vah = stock.get("vah_60d")
            if poc or val or vah:
                volume_text = f"""
                <b>Volume Profile (60d):</b> POC ${poc:,.2f} | VAL ${val:,.2f} - VAH ${vah:,.2f}
                """
                elements.append(Paragraph(volume_text, styles["Normal"]))

            elements.append(Spacer(1, 0.15 * inch))
    else:
        elements.append(Paragraph("BUY SETUPS (0 stocks)", section_style))
        elements.append(
            Paragraph("No quality buy signals at this time.", styles["Normal"])
        )
        elements.append(PageBreak())

    # === SECTION 2: SELL SETUPS ===
    if not sell_df.empty:
        elements.append(PageBreak())
        elements.append(
            Paragraph(f"SELL SETUPS ({len(sell_df)} stocks)", section_style)
        )
        elements.append(Spacer(1, 0.1 * inch))

        # Sort by vol_rr if available, otherwise by ticker
        sort_col = "vol_rr" if "vol_rr" in sell_df.columns else "ticker"
        sort_asc = False if sort_col == "vol_rr" else True

        for rank, (_, stock) in enumerate(
            sell_df.sort_values(sort_col, ascending=sort_asc).iterrows(), 1
        ):
            ticker = stock["ticker"]
            price = stock["current_price"]
            signal = stock["signal"]
            r1_quality = stock.get("r1_quality", "N/A")
            r2_quality = stock.get("r2_quality", "N/A")
            r3_quality = stock.get("r3_quality", "N/A")
            rsi = stock.get("rsi")

            # Short entry quality assessment
            short_entry_quality = stock.get("short_entry_quality", "N/A")
            short_entry_flag = stock.get("short_entry_flag", "")
            short_entry_note = stock.get("short_entry_note", "")
            vol_rr = stock.get("vol_rr", 0)

            # Color code the R:R ratio
            if vol_rr >= 3:
                rr_color = "#27ae60"  # Green
            elif vol_rr >= 2:
                rr_color = "#3498db"  # Blue
            elif vol_rr >= 1:
                rr_color = "#f39c12"  # Orange
            else:
                rr_color = "#e74c3c"  # Red

            # Stock header with ranking, short entry quality and flag
            short_flag_text = f" {short_entry_flag}" if short_entry_flag else ""
            header_text = f"<b>#{rank}. {ticker} - ${price:,.2f}</b> | <font color='#2980b9'><b>{short_entry_quality}{short_flag_text}</b></font> | Vol R:R: <font color='{rr_color}'><b>{vol_rr:.1f}:1</b></font>"
            elements.append(Paragraph(header_text, subsection_style))

            # RSI and Signal line
            rsi_text = f"{rsi:.1f}" if rsi else "N/A"
            signal_line = f"<b>RSI:</b> {rsi_text} | <b>Signal:</b> {signal}"
            elements.append(Paragraph(signal_line, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Volatility Stop
            suggested_stop = stock.get("suggested_stop_pct", 5.0)
            vol_stop_price = price * (
                1 + suggested_stop / 100
            )  # For shorts, stop is above

            stop_info = f"<b>Volatility Stop ({suggested_stop:.1f}%):</b> ${vol_stop_price:.2f} loss = <font color='#e74c3c'>-{suggested_stop:.1f}%</font>"
            elements.append(Paragraph(stop_info, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Target S1
            s1 = stock.get("s1", 0)
            if s1 and price > 0 and s1 < price:
                reward_pct = (price - s1) / price * 100
                target_text = f"<b>Target S1:</b> ${s1:.2f} gain = <font color='#27ae60'>+{reward_pct:.1f}%</font>"
                elements.append(Paragraph(target_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Resistance R1 and Volatility info
            r1 = stock.get("r1", 0)
            volatility_class = stock.get("volatility_class", "")
            daily_range = stock.get("avg_daily_range_pct", 0)
            if r1 and price > 0:
                distance_to_r1 = abs((r1 - price) / price * 100)
                r1_text = f"<b>Resistance R1:</b> ${r1:.2f} ({distance_to_r1:.1f}% away) | <b>Volatility:</b> {volatility_class} (~{daily_range:.2f}% daily range)"
                elements.append(Paragraph(r1_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Entry note
            if short_entry_note:
                elements.append(
                    Paragraph(f"<i>{short_entry_note}</i>", styles["Normal"])
                )
            elements.append(Spacer(1, 0.05 * inch))

            # Support levels (downside targets for shorts)
            s1 = stock.get("s1")
            s2 = stock.get("s2")
            s3 = stock.get("s3")

            exit_text = "<b>Downside Targets (Support Levels):</b><br/>"
            if s1:
                gain1 = (price - s1) / price * 100 if price > 0 and s1 < price else 0
                exit_text += f"- S1: ${s1:,.2f} ({gain1:.1f}% gain)<br/>"
            if s2:
                gain2 = (price - s2) / price * 100 if price > 0 and s2 < price else 0
                exit_text += f"- S2: ${s2:,.2f} ({gain2:.1f}% gain)<br/>"
            if s3:
                gain3 = (price - s3) / price * 100 if price > 0 and s3 < price else 0
                exit_text += f"- S3: ${s3:,.2f} ({gain3:.1f}% gain)<br/>"
            elements.append(Paragraph(exit_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # S1/S2/S3 quality assessments (matching portfolio format)
            s1_quality = stock.get("s1_quality", "N/A")
            s1_quality_note = stock.get("s1_quality_note", "")
            s2_quality = stock.get("s2_quality", "N/A")
            s2_quality_note = stock.get("s2_quality_note", "")
            s3_quality = stock.get("s3_quality", "N/A")
            s3_quality_note = stock.get("s3_quality_note", "")

            quality_text = f"""
            <b>Cover Quality S1:</b> {s1_quality} - {s1_quality_note}<br/>
            <b>Cover Quality S2:</b> {s2_quality} - {s2_quality_note}<br/>
            <b>Cover Quality S3:</b> {s3_quality} - {s3_quality_note}
            """
            elements.append(Paragraph(quality_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Moving averages
            d50 = stock.get("d50")
            d100 = stock.get("d100")
            d200 = stock.get("d200")
            if d50 or d100 or d200:
                ma_text = "<b>Moving Averages:</b> "
                ma_parts = []
                if d50:
                    ma_parts.append(f"D50 ${d50:,.2f}")
                if d100:
                    ma_parts.append(f"D100 ${d100:,.2f}")
                if d200:
                    ma_parts.append(f"D200 ${d200:,.2f}")
                ma_text += " | ".join(ma_parts)
                elements.append(Paragraph(ma_text, styles["Normal"]))

            elements.append(Spacer(1, 0.15 * inch))
    else:
        elements.append(PageBreak())
        elements.append(Paragraph("SELL SETUPS (0 stocks)", section_style))
        elements.append(Paragraph("No bearish signals at this time.", styles["Normal"]))

    # Glossary Section
    elements.append(PageBreak())
    elements.append(Paragraph("GLOSSARY OF TERMS", section_style))
    elements.append(Spacer(1, 0.1 * inch))

    glossary_items = [
        (
            "<b>Support Levels (S1, S2, S3):</b>",
            "Price levels where buying pressure may prevent further decline. S1 is the nearest support.",
        ),
        (
            "<b>Resistance Levels (R1, R2, R3):</b>",
            "Price levels where selling pressure may prevent further advance. R1 is the nearest resistance.",
        ),
        (
            "<b>RSI (Relative Strength Index):</b>",
            "Momentum indicator (0-100). Below 30 = oversold, above 70 = overbought.",
        ),
        (
            "<b>Volatility Stop:</b>",
            "Dynamic stop-loss based on stock's volatility (typically 5-8% below entry for buys).",
        ),
        (
            "<b>Vol R:R (Volatility Risk-Reward):</b>",
            "Ratio of potential reward to volatility stop risk. Higher is better (3+ is excellent).",
        ),
        (
            "<b>D50/D100/D200:</b>",
            "50/100/200-day moving averages. Key trend indicators and support/resistance levels.",
        ),
        (
            "<b>POC (Point of Control):</b>",
            "Price level with highest trading volume. Strong support/resistance.",
        ),
        (
            "<b>VAH/VAL (Value Area High/Low):</b>",
            "Price range containing 70% of volume. Defines fair value zone.",
        ),
        (
            "<b>HVN (High Volume Node):</b>",
            "Price level with significant trading activity. Acts as support/resistance.",
        ),
        (
            "<b>FULL HOLD + ADD:</b>",
            "Strong bullish signal. Stock is in uptrend with multiple support levels.",
        ),
        (
            "<b>FULL SELL + SHORT:</b>",
            "Strong bearish signal. Stock is in downtrend with multiple resistance levels.",
        ),
    ]

    for term, definition in glossary_items:
        elements.append(Paragraph(f"{term} {definition}", styles["Normal"]))
        elements.append(Spacer(1, 0.08 * inch))

    # Build PDF
    doc.build(elements)
    print(f"[OK] PDF report created: {output_file}")


def create_best_trades_pdf(buy_df, sell_df, output_file, category="", regime=None):
    """
    Create PDF with ranked best trading opportunities

    Args:
        buy_df: DataFrame with buy opportunities
        sell_df: DataFrame with sell opportunities
        output_file: Path to save PDF file
        category: Label (e.g., 'S&P 500', 'NASDAQ 100')
        regime: Market regime dict from analyze_market_regime() (optional, for display)
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    doc = SimpleDocTemplate(
        str(output_file),
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#2c3e50"),
        spaceAfter=20,
        alignment=TA_CENTER,
    )

    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=16,
        textColor=colors.HexColor("#2980b9"),
        spaceAfter=12,
        spaceBefore=20,
    )

    # === TITLE PAGE ===
    report_title = category.replace("&", "&amp;")
    cover_title = f"Best Trading Setups<br/>{report_title}<br/><font size=14>{datetime.now().strftime('%B %d, %Y')}</font>"
    elements.append(Paragraph(cover_title, title_style))
    elements.append(Spacer(1, 0.3 * inch))

    # Market Regime Banner (if provided)
    if regime:
        regime_color = {
            "GREEN": "#27ae60",  # green
            "YELLOW": "#f39c12",  # yellow/orange
            "ORANGE": "#e67e22",  # orange
            "RED": "#c0392b",  # red
        }.get(
            regime["tier"], "#95a5a6"
        )  # gray default

        regime_banner = f"""
        <para align=center bgcolor='{regime_color}' textColor='white'>
        <b><font size=14>{regime['emoji']} MARKET REGIME: {regime['tier']}</font></b><br/>
        <font size=11>QQQ: {regime['signal']} | {regime['weekly_state']} / {regime['daily_state']}</font><br/>
        <font size=10>{regime['message']}</font>
        </para>
        """
        elements.append(Paragraph(regime_banner, styles["Normal"]))
        elements.append(Spacer(1, 0.2 * inch))

    # Calculate Vol R:R and rank - OPTION 1 approach
    buy_scored = buy_df.copy() if not buy_df.empty else pd.DataFrame()
    sell_scored = sell_df.copy() if not sell_df.empty else pd.DataFrame()

    if not buy_scored.empty:
        # Filter for EXCELLENT and GOOD entry quality only (stop-aware)
        quality_col = (
            "entry_quality" if "entry_quality" in buy_scored.columns else "buy_quality"
        )
        buy_scored = buy_scored[
            buy_scored[quality_col].isin(["EXCELLENT", "GOOD"])
        ].copy()

        if not buy_scored.empty:
            # Calculate Vol R:R for ranking
            def calc_vol_rr(row):
                price = row.get("current_price", 0)
                r1 = row.get("r1", 0)
                suggested_stop = row.get("suggested_stop_pct", 5.0)

                reward = (
                    ((r1 - price) / price * 100)
                    if price > 0 and r1 > 0 and r1 > price
                    else 0
                )
                return (
                    reward / suggested_stop if suggested_stop > 0 and reward > 0 else 0
                )

            buy_scored["vol_rr"] = buy_scored.apply(calc_vol_rr, axis=1)

            # Sort by Vol R:R (highest first), take top 15
            buy_scored = buy_scored.sort_values("vol_rr", ascending=False).head(15)

    if not sell_scored.empty:
        # Filter for EXCELLENT and GOOD short entry quality
        quality_col = (
            "short_entry_quality"
            if "short_entry_quality" in sell_scored.columns
            else "r1_quality"
        )
        sell_scored = sell_scored[
            sell_scored[quality_col].isin(["EXCELLENT", "GOOD"])
        ].copy()

        if not sell_scored.empty:
            # Calculate Vol R:R for ranking
            def calc_vol_rr(row):
                price = row.get("current_price", 0)
                s1 = row.get("s1", 0)
                suggested_stop = row.get("suggested_stop_pct", 5.0)

                reward = (
                    ((price - s1) / price * 100)
                    if price > 0 and s1 > 0 and price > s1
                    else 0
                )
                return (
                    reward / suggested_stop if suggested_stop > 0 and reward > 0 else 0
                )

            sell_scored["vol_rr"] = sell_scored.apply(calc_vol_rr, axis=1)

            # Sort by Vol R:R (highest first), take top 15
            sell_scored = sell_scored.sort_values("vol_rr", ascending=False).head(15)

    # Summary
    summary_text = f"""
    <b>Best Trading Setups - Volatility-Based Ranking</b><br/>
    <br/>
    <b>Buy Setups:</b> {len(buy_scored)} EXCELLENT/GOOD quality (stop-aware) setups<br/>
    <b>Sell Setups:</b> {len(sell_scored)} EXCELLENT/GOOD quality (stop-aware) setups<br/>
    <br/>
    <b>Selection Criteria:</b><br/>
    • <b>Quality Filter (Buys):</b> Only EXCELLENT or GOOD quality (supports within 8% stop tolerance)<br/>
    • <b>Quality Filter (Shorts):</b> Only EXCELLENT or GOOD quality (resistances above 8% stop tolerance)<br/>
    • <b>Ranking:</b> Sorted by Vol R:R (Volatility Risk/Reward Ratio) - highest first<br/>
    <br/>
    <b>What is Vol R:R?</b><br/>
    The ratio of potential reward (to Target R1/S1) divided by volatility-based stop loss %.<br/>
    Example: +12% reward / 6% vol stop = 2:1 Vol R:R<br/>
    <br/>
    <b>Quality Ratings (Stop-Aware):</b><br/>
    • <b>BUY ENTRIES:</b> Multiple strong supports within your 8% stop range protect long positions<br/>
    • <b>SHORT ENTRIES:</b> Multiple strong resistances above your 8% stop protect short positions<br/>
    • Filters out entries where good supports/resistances are beyond your stop tolerance<br/>
    <br/>
    <i>Focus on higher Vol R:R setups (2:1 or better) for optimal risk-adjusted returns.</i>
    """
    elements.append(Paragraph(summary_text, styles["Normal"]))

    # === GLOSSARY OF TERMS (moved to start) ===
    elements.append(Paragraph("GLOSSARY OF TERMS", section_style))

    glossary_items = [
        ("<b>Quality Ratings:</b>", ""),
        (
            "  • <b>EXCELLENT:</b>",
            "2+ strong supports accessible within stop tolerance. Safest entries.",
        ),
        (
            "  • <b>GOOD:</b>",
            "1+ good support accessible within stop tolerance. Solid risk management.",
        ),
        (
            "  • <b>OK:</b>",
            "Moderate supports available. Acceptable but watch closely.",
        ),
        (
            "  • <b>CAUTION:</b>",
            "Thin or extended. Limited protection within stop range.",
        ),
        ("", ""),
        ("<b>Quality Flags:</b>", ""),
        (
            "  • <b>SAFE ENTRY:</b>",
            "Multiple excellent supports protect your stop level.",
        ),
        ("  • <b>IDEAL:</b>", "Near major support with good upside potential."),
        ("  • <b>ACCEPTABLE:</b>", "Reasonable entry but monitor risk carefully."),
        ("  • <b>THIN:</b>", "Limited support structure. Higher risk."),
        ("  • <b>EXTENDED:</b>", "Far from supports. Consider waiting for pullback."),
        ("  • <b>WAIT:</b>", "No accessible supports within 8% stop. Do not enter."),
        ("", ""),
        (
            "<b>Vol R:R (Volatility Risk-Reward):</b>",
            "Ratio of potential gain to volatility stop risk. 3+ = excellent, 2+ = good, 1+ = acceptable.",
        ),
        (
            "<b>Volatility Stop:</b>",
            "Stop-loss based on stock's price movement (typically 5-8%). Limits maximum loss.",
        ),
        (
            "<b>Stop Tolerance:</b>",
            "Maximum acceptable stop distance (8%). Entries are rated based on supports within this range.",
        ),
        (
            "<b>RSI:</b>",
            "Relative Strength Index (0-100). <30 oversold, >70 overbought.",
        ),
        ("<b>S1/R1:</b>", "Primary support and resistance levels. Key price targets."),
        (
            "<b>D50/D100/D200:</b>",
            "50/100/200-day moving averages. Trend and support indicators.",
        ),
    ]

    for term, definition in glossary_items:
        if term and definition:
            elements.append(Paragraph(f"{term} {definition}", styles["Normal"]))
            elements.append(Spacer(1, 0.04 * inch))
        elif term and not definition:
            elements.append(Paragraph(term, styles["Normal"]))
            elements.append(Spacer(1, 0.03 * inch))
        else:
            elements.append(Spacer(1, 0.06 * inch))

    elements.append(PageBreak())

    # === TOP BUY SETUPS ===
    if not buy_scored.empty:
        elements.append(Paragraph(f"TOP {len(buy_scored)} BUY SETUPS", section_style))
        elements.append(Spacer(1, 0.1 * inch))

        for rank, (_, stock) in enumerate(buy_scored.iterrows(), 1):
            ticker = stock["ticker"]
            price = stock.get("current_price", 0)
            # Use entry_quality (stop-aware) if available
            quality = stock.get("entry_quality", stock.get("buy_quality", ""))
            entry_flag = stock.get("entry_flag", "")
            signal = stock.get("signal", "")
            s1 = stock.get("s1", 0)
            r1 = stock.get("r1", 0)
            rsi = stock.get("rsi")
            vol_rr = stock.get("vol_rr", 0)

            # Calculate reward potential
            reward_pct = (
                ((r1 - price) / price * 100)
                if price > 0 and r1 > 0 and r1 > price
                else 0
            )
            distance_to_s1 = (
                abs((price - s1) / price * 100) if price > 0 and s1 > 0 else 0
            )

            # Volatility-based stop
            suggested_stop = stock.get("suggested_stop_pct", 5.0)
            vol_stop_price = price * (1 - suggested_stop / 100)

            # Quality badge color
            if quality == "EXCELLENT":
                quality_color = "#27ae60"  # Green
            else:  # GOOD
                quality_color = "#3498db"  # Blue

            # Vol R:R color
            if vol_rr >= 3:
                rr_color = "#27ae60"  # Green
            elif vol_rr >= 2:
                rr_color = "#3498db"  # Blue
            elif vol_rr >= 1:
                rr_color = "#f39c12"  # Orange
            else:
                rr_color = "#e74c3c"  # Red

            # Volatility class display
            volatility_class = stock.get("volatility_class", "")
            daily_range = stock.get("avg_daily_range_pct", 0)

            # Stop-aware details
            stop_tolerance = stock.get("stop_tolerance_pct", 8.0)
            stop_level = stock.get("stop_level", price * 0.92)
            accessible_supports = stock.get("accessible_supports_count", 0)
            entry_note = stock.get("entry_note", "")

            # Format entry flag for display
            entry_flag_display = f" {entry_flag}" if entry_flag else ""

            card_text = f"""
            <font size=12><b>#{rank}. {ticker} - ${price:.2f}</b> | <font color="{quality_color}"><b>{quality}</b>{entry_flag_display}</font> | Vol R:R: <font color="{rr_color}"><b>{vol_rr:.1f}:1</b></font></font><br/>
            <b>RSI:</b> {f'{rsi:.1f}' if rsi else 'N/A'} | <b>Signal:</b> {signal or 'N/A'}<br/>
            <br/>
            <b>Volatility Stop ({suggested_stop:.1f}%):</b> ${vol_stop_price:.2f} loss = <font color="#e74c3c">-{suggested_stop:.1f}%</font><br/>
            <b>8% Stop Tolerance:</b> ${stop_level:.2f} | <b>Accessible Supports:</b> {accessible_supports} within range<br/>
            <b>Target R1:</b> ${r1:.2f} gain = <font color="#27ae60">+{reward_pct:.1f}%</font><br/>
            <br/>
            <b>Support S1:</b> ${s1:.2f} ({distance_to_s1:.1f}% away) | <b>Volatility:</b> {volatility_class} (~{daily_range:.2f}% daily range)<br/>
            <i>{entry_note}</i>
            """
            elements.append(Paragraph(card_text, styles["Normal"]))
            elements.append(Spacer(1, 0.15 * inch))
    else:
        elements.append(Paragraph("TOP BUY SETUPS", section_style))
        elements.append(
            Paragraph(
                "No EXCELLENT/GOOD entry quality (stop-aware) buy opportunities found. Good supports may exist but are beyond 8% stop tolerance.",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 0.2 * inch))

    # === TOP SELL SETUPS ===
    if not sell_scored.empty:
        elements.append(Paragraph(f"TOP {len(sell_scored)} SELL SETUPS", section_style))
        elements.append(Spacer(1, 0.1 * inch))

        for rank, (_, stock) in enumerate(sell_scored.iterrows(), 1):
            ticker = stock["ticker"]
            price = stock.get("current_price", 0)
            quality = stock.get("short_entry_quality", stock.get("r1_quality", ""))
            short_entry_flag = stock.get("short_entry_flag", "")
            signal = stock.get("signal", "")
            r1 = stock.get("r1", 0)
            s1 = stock.get("s1", 0)
            rsi = stock.get("rsi")
            vol_rr = stock.get("vol_rr", 0)

            # Calculate reward potential
            reward_pct = (
                ((price - s1) / price * 100)
                if price > 0 and s1 > 0 and price > s1
                else 0
            )
            distance_to_r1 = (
                abs((r1 - price) / price * 100) if price > 0 and r1 > 0 else 0
            )

            # Volatility-based stop (for sell: stop is above price)
            suggested_stop = stock.get("suggested_stop_pct", 5.0)
            vol_stop_price = price * (1 + suggested_stop / 100)

            # Quality badge color
            if quality == "EXCELLENT":
                quality_color = "#27ae60"  # Green
            else:  # GOOD
                quality_color = "#3498db"  # Blue

            # Vol R:R color
            if vol_rr >= 3:
                rr_color = "#27ae60"  # Green
            elif vol_rr >= 2:
                rr_color = "#3498db"  # Blue
            elif vol_rr >= 1:
                rr_color = "#f39c12"  # Orange
            else:
                rr_color = "#e74c3c"  # Red

            # Volatility class display
            volatility_class = stock.get("volatility_class", "")
            daily_range = stock.get("avg_daily_range_pct", 0)

            # Format short entry flag for display
            short_flag_text = f" {short_entry_flag}" if short_entry_flag else ""

            card_text = f"""
            <font size=12><b>#{rank}. {ticker} - ${price:.2f}</b> | <font color="{quality_color}"><b>{quality}{short_flag_text}</b></font> | Vol R:R: <font color="{rr_color}"><b>{vol_rr:.1f}:1</b></font></font><br/>
            <b>RSI:</b> {f'{rsi:.1f}' if rsi else 'N/A'} | <b>Signal:</b> {signal or 'N/A'}<br/>
            <br/>
            <b>Volatility Stop ({suggested_stop:.1f}%):</b> ${vol_stop_price:.2f} loss = <font color="#e74c3c">-{suggested_stop:.1f}%</font><br/>
            <b>Target S1:</b> ${s1:.2f} gain = <font color="#27ae60">+{reward_pct:.1f}%</font><br/>
            <br/>
            <b>Resistance R1:</b> ${r1:.2f} ({distance_to_r1:.1f}% away) | <b>Volatility:</b> {volatility_class} (~{daily_range:.2f}% daily range)
            """
            elements.append(Paragraph(card_text, styles["Normal"]))
            elements.append(Spacer(1, 0.15 * inch))
    else:
        elements.append(Paragraph("TOP SELL SETUPS", section_style))
        elements.append(
            Paragraph(
                "No EXCELLENT/GOOD quality sell opportunities found.", styles["Normal"]
            )
        )

    # Build PDF
    doc.build(elements)
    print(f"[OK] Best Trades PDF created: {output_file}")


if __name__ == "__main__":
    import argparse
    from pathlib import Path
    import sys

    # Fix Windows console encoding for emojis
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except:
            pass  # Fallback to default encoding

    parser = argparse.ArgumentParser(
        description="Scan stocks for FULL HOLD + ADD signals"
    )
    parser.add_argument(
        "--daily-bars", type=int, default=60, help="Daily bars (default: 60)"
    )
    parser.add_argument(
        "--weekly-bars", type=int, default=52, help="Weekly bars (default: 52)"
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=2,
        help="Threads (default: 2, higher values risk rate limiting)",
    )
    parser.add_argument(
        "--as-of-date",
        type=str,
        default=None,
        help="Historical date for simulation (YYYY-MM-DD format)",
    )

    args = parser.parse_args()

    # Setup results directory
    results_dir = Path.cwd() / "scanner_results"
    results_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    overall_start = datetime.now()

    # === REGIME ANALYSIS (MANDATORY) ===
    regime = analyze_market_regime(
        daily_bars=args.daily_bars,
        weekly_bars=args.weekly_bars,
        as_of_date=args.as_of_date,
    )

    # === SCAN 1: S&P 500 ===
    print("\n" + "=" * 80)
    print("SCAN 1 of 3: S&P 500")
    print("=" * 80 + "\n")

    print("Fetching S&P 500 ticker list...")
    sp500_tickers = get_sp500_tickers()
    print(f"Found {len(sp500_tickers)} S&P 500 stocks\n")

    start_time = datetime.now()
    sp500_results = scan_stocks(
        sp500_tickers,
        category="S&P 500",
        daily_bars=args.daily_bars,
        weekly_bars=args.weekly_bars,
        concurrency=args.concurrency,
        as_of_date=args.as_of_date,
        regime=regime,
    )
    sp500_elapsed = (datetime.now() - start_time).total_seconds()

    if not sp500_results.empty:
        sp500_buy = filter_buy_signals(sp500_results, "FULL HOLD + ADD", regime=regime)
        sp500_sell = filter_sell_signals(sp500_results)
        print(f"\n🎯 S&P 500: {len(sp500_buy)} FULL HOLD + ADD signals found")
        print(f"⚠️ S&P 500: {len(sp500_sell)} bearish signals found")

        if not sp500_buy.empty or not sp500_sell.empty:
            # Save S&P 500 results
            xlsx_path = results_dir / f"sp500_analysis_{timestamp}.xlsx"
            pdf_path = results_dir / f"scanner_report_sp500_{timestamp}.pdf"

            create_excel_output(sp500_buy, sp500_sell, xlsx_path, category="S&P 500")
            create_pdf_report(
                sp500_buy, sp500_results, pdf_path, timestamp, category="S&P 500"
            )

            print(f"  âœ“ Excel: {xlsx_path.name}")
            print(f"  âœ“ PDF: {pdf_path.name}")

            # Create Best Trades reports
            best_trades_xlsx = (
                results_dir / "sp500" / f"sp500_best_trades_{timestamp}.xlsx"
            )
            best_trades_pdf = (
                results_dir / "sp500" / f"sp500_best_trades_{timestamp}.pdf"
            )
            (results_dir / "sp500").mkdir(exist_ok=True)
            create_best_trades_excel(
                sp500_buy, sp500_sell, best_trades_xlsx, category="S&P 500"
            )
            create_best_trades_pdf(
                sp500_buy, sp500_sell, best_trades_pdf, category="S&P 500"
            )
            print(f"  Best Trades: {best_trades_xlsx.name}, {best_trades_pdf.name}")

    # === SCAN 2: NASDAQ 100 ===
    print("\n" + "=" * 80)
    print("SCAN 2 of 3: NASDAQ 100")
    print("=" * 80 + "\n")

    print("Fetching NASDAQ 100 ticker list...")
    nasdaq100_tickers = get_nasdaq100_tickers()
    print(f"Found {len(nasdaq100_tickers)} NASDAQ 100 stocks\n")

    start_time = datetime.now()
    nasdaq100_results = scan_stocks(
        nasdaq100_tickers,
        category="NASDAQ 100",
        daily_bars=args.daily_bars,
        weekly_bars=args.weekly_bars,
        concurrency=args.concurrency,
        as_of_date=args.as_of_date,
        regime=regime,
    )
    nasdaq100_elapsed = (datetime.now() - start_time).total_seconds()

    if not nasdaq100_results.empty:
        nasdaq100_buy = filter_buy_signals(
            nasdaq100_results, "FULL HOLD + ADD", regime=regime
        )
        nasdaq100_sell = filter_sell_signals(nasdaq100_results)
        print(f"\nðŸŽ¯ NASDAQ 100: {len(nasdaq100_buy)} FULL HOLD + ADD signals found")
        print(f"âš ï¸ NASDAQ 100: {len(nasdaq100_sell)} bearish signals found")

        if not nasdaq100_buy.empty or not nasdaq100_sell.empty:
            # Save NASDAQ 100 results
            xlsx_path = results_dir / f"nasdaq100_analysis_{timestamp}.xlsx"
            pdf_path = results_dir / f"scanner_report_nasdaq100_{timestamp}.pdf"

            create_excel_output(
                nasdaq100_buy, nasdaq100_sell, xlsx_path, category="NASDAQ 100"
            )
            create_pdf_report(
                nasdaq100_buy,
                nasdaq100_results,
                pdf_path,
                timestamp,
                category="NASDAQ 100",
                regime=regime,
            )

            print(f"  âœ“ Excel: {xlsx_path.name}")
            best_trades_xlsx = (
                results_dir / "nasdaq100" / f"nasdaq100_best_trades_{timestamp}.xlsx"
            )
            best_trades_pdf = (
                results_dir / "nasdaq100" / f"nasdaq100_best_trades_{timestamp}.pdf"
            )
            (results_dir / "nasdaq100").mkdir(exist_ok=True)
            create_best_trades_excel(
                nasdaq100_buy, nasdaq100_sell, best_trades_xlsx, category="NASDAQ 100"
            )
            create_best_trades_pdf(
                nasdaq100_buy,
                nasdaq100_sell,
                best_trades_pdf,
                category="NASDAQ 100",
                regime=regime,
            )

            print(f"  âœ“ PDF: {pdf_path.name}")

            print(f"  Best Trades: {best_trades_xlsx.name}, {best_trades_pdf.name}")
    # === SCAN 3: PORTFOLIO STOCKS ===
    print("\n" + "=" * 80)
    print("SCAN 3 of 3: PORTFOLIO STOCKS")
    print("=" * 80 + "\n")

    print("Loading portfolio tickers from stocks.txt...")
    portfolio_tickers = get_portfolio_tickers()

    if portfolio_tickers:
        print(f"Found {len(portfolio_tickers)} portfolio stocks\n")

        start_time = datetime.now()
        portfolio_results = scan_stocks(
            portfolio_tickers,
            category="Portfolio",
            daily_bars=args.daily_bars,
            weekly_bars=args.weekly_bars,
            concurrency=args.concurrency,
            as_of_date=args.as_of_date,
            regime=regime,
        )
        portfolio_elapsed = (datetime.now() - start_time).total_seconds()

        if not portfolio_results.empty:
            # For portfolio, include ALL stocks (not just FULL HOLD + ADD)
            # But still create separate sheets by signal type
            print(
                f"\nðŸŽ¯ Portfolio: {len(portfolio_results)} stocks scanned (all included)"
            )

            # Count FULL HOLD + ADD for display
            portfolio_buy = filter_buy_signals(
                portfolio_results, "FULL HOLD + ADD", regime=regime
            )
            print(f"  - {len(portfolio_buy)} with FULL HOLD + ADD signal")

            # Save ALL portfolio results (not filtered)
            xlsx_path = results_dir / f"portfolio_scanner_{timestamp}.xlsx"
            pdf_path = results_dir / f"scanner_report_portfolio_{timestamp}.pdf"

            # Pass ALL results for portfolio
            create_portfolio_excel(portfolio_results, xlsx_path, category="Portfolio")

            # For PDF, still show FULL HOLD + ADD in detail (with all stocks in summary)
            if not portfolio_buy.empty:
                create_pdf_report(
                    portfolio_buy,
                    portfolio_results,
                    pdf_path,
                    timestamp,
                    category="Portfolio",
                    regime=regime,
                )

            print(f"  âœ“ Excel: {xlsx_path.name}")
            print(f"  âœ“ PDF: {pdf_path.name}")
    else:
        print("âš ï¸  No portfolio tickers found in stocks.txt")

    # === FINAL SUMMARY ===
    total_elapsed = (datetime.now() - overall_start).total_seconds()

    print("\n" + "=" * 80)
    print("SCAN COMPLETE - SUMMARY")
    print("=" * 80)
    print(f"[TIME] Total time: {total_elapsed:.1f}s")
    print(f"\nðŸ“Š Results by Category:")

    if not sp500_results.empty:
        sp500_buy_count = len(
            filter_buy_signals(sp500_results, "FULL HOLD + ADD", regime=regime)
        )
        print(
            f"  S&P 500:     {sp500_buy_count:3d} FULL HOLD + ADD signals ({sp500_elapsed:.1f}s)"
        )

    if not nasdaq100_results.empty:
        nasdaq100_buy_count = len(
            filter_buy_signals(nasdaq100_results, "FULL HOLD + ADD", regime=regime)
        )
        print(
            f"  NASDAQ 100:  {nasdaq100_buy_count:3d} FULL HOLD + ADD signals ({nasdaq100_elapsed:.1f}s)"
        )

    if portfolio_tickers and not portfolio_results.empty:
        portfolio_buy_count = len(
            filter_buy_signals(portfolio_results, "FULL HOLD + ADD", regime=regime)
        )
        print(
            f"  Portfolio:   {portfolio_buy_count:3d} FULL HOLD + ADD signals ({portfolio_elapsed:.1f}s)"
        )

    # Cleanup old scans (keep 1 most recent per category)
    print("\nðŸ“ Managing scan history...")
    cleanup_old_scans(results_dir, max_files=1)
