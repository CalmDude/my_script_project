"""
Portfolio Report Generation - Phase 3
Creates Trading Playbook PDF and Portfolio Tracker Excel
"""

from pathlib import Path
from datetime import datetime
import pandas as pd
import logging
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

# Configure logging
logger = logging.getLogger(__name__)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def create_trading_playbook_pdf(portfolio_data, output_path, timestamp_str):
    """
    Create comprehensive Trading Playbook PDF.

    Args:
        portfolio_data: dict with keys:
            - 'positions': list of position dicts with analysis results
            - 'portfolio_total': total portfolio value
            - 'summary': dict with counts and totals
        output_path: Path to save PDF
        timestamp_str: Timestamp string for header

    Each position dict should have:
        - ticker, signal, current_price
        - target_pct, current_value, gap_value
        - s1, s2, s3, r1, r2, r3
        - d50, d100, d200
        - buy_quality, buy_tranches, sell_tranches
        - action (BUY/SELL/HOLD/WAIT)
    """
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#2c3e50"),
        spaceAfter=20,
        alignment=TA_CENTER,
    )

    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=16,
        textColor=colors.HexColor("#2980b9"),
        spaceAfter=12,
        spaceBefore=20,
    )

    subsection_style = ParagraphStyle(
        "Subsection",
        parent=styles["Heading3"],
        fontSize=12,
        textColor=colors.HexColor("#34495e"),
        spaceAfter=6,
        spaceBefore=10,
    )

    # === PAGE 1: DASHBOARD ===
    cover_title = f"Portfolio Playbook<br/><font size=14>{datetime.now().strftime('%B %d, %Y')}</font>"
    elements.append(Paragraph(cover_title, title_style))
    elements.append(Spacer(1, 0.3 * inch))

    # Portfolio summary
    summary = portfolio_data.get("summary", {})
    # Count only positions with actual holdings
    active_positions = sum(
        1 for p in portfolio_data["positions"] if p.get("current_value", 0) > 0
    )
    buy_count = summary.get("buy_count", 0)
    sell_count = summary.get("sell_count", 0)
    hold_count = summary.get("hold_count", 0)

    # Portfolio value breakdown
    total_value = summary.get(
        "total_portfolio_value", portfolio_data["portfolio_total"]
    )
    current_holdings = portfolio_data["portfolio_total"]
    cash_available = summary.get("cash_available", total_value - current_holdings)
    cash_pct = (cash_available / total_value * 100) if total_value > 0 else 0

    dashboard_text = f"""
    <b>Portfolio Snapshot</b><br/>
    - Total Portfolio Value: ${total_value:,.0f}<br/>
    - Current Holdings: ${current_holdings:,.0f}<br/>
    - Cash Available: ${cash_available:,.0f} (~{cash_pct:.0f}% dry powder)<br/>
    - Active Positions: {active_positions}<br/>
    <br/>
    <b>Today's Actions</b><br/>
    - {buy_count} BUY opportunities (FULL HOLD + ADD)<br/>
    - {sell_count} SELL actions (reduction signals)<br/>
    - {hold_count} HOLD positions (no action)<br/>
    <br/>
    <b>Report Sections</b><br/>
    1. Buy Actions - Add to underweight positions<br/>
    2. Sell Actions - Reduce overweight or bearish positions<br/>
    3. Hold Positions - At target or waiting for signal<br/>
    """
    elements.append(Paragraph(dashboard_text, styles["Normal"]))
    elements.append(PageBreak())

    # === SECTION 2: BUY ACTIONS ===
    # Only show buy actions for FULL HOLD + ADD signal
    buy_positions = [
        p
        for p in portfolio_data["positions"]
        if p.get("action") == "BUY" and p["signal"] == "FULL HOLD + ADD"
    ]

    if buy_positions:
        elements.append(
            Paragraph(f"BUY ACTIONS ({len(buy_positions)} positions)", section_style)
        )
        elements.append(Spacer(1, 0.1 * inch))

        for pos in buy_positions:
            ticker = pos["ticker"]
            signal = pos["signal"]
            price = pos["current_price"]
            target_pct = pos.get("target_pct", 0) * 100
            current_value = pos.get("current_value", 0)
            gap_value = pos.get("gap_value", 0)

            # Position header
            header_text = f"<b>{ticker}</b> - ${price:,.2f} ({signal})"
            elements.append(Paragraph(header_text, subsection_style))

            # Allocation info
            alloc_text = f"""
            <b>Allocation:</b> Target {target_pct:.0f}% | Current ${current_value:,.0f} | Need ${gap_value:,.0f}
            """
            elements.append(Paragraph(alloc_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Buy tranches (simplified - no quality notes repeated)
            tranches = pos.get("buy_tranches", [])
            if tranches:
                elements.append(Paragraph("<b>Buy Plan:</b>", styles["Normal"]))
                # Reverse order to show S1 first (shallowest/most likely), then S2, then S3
                for tranche in reversed(tranches):
                    # Handle both old format (4 items) and new format (6 items with quality)
                    if len(tranche) == 6:
                        (
                            price_level,
                            amount,
                            level_name,
                            status,
                            quality,
                            quality_note,
                        ) = tranche
                        quantity = (
                            amount / price_level
                            if price_level and price_level > 0
                            else 0
                        )
                        if amount > 0:
                            tranche_text = f"- {level_name} at ${price_level:,.2f}: Buy {quantity:.2f} shares (${amount:,.0f})"
                        else:
                            tranche_text = f"- {level_name}: {status}"
                    else:
                        # Old format fallback
                        price_level, amount, level_name, status = tranche[:4]
                        quantity = (
                            amount / price_level
                            if price_level and price_level > 0
                            else 0
                        )
                        if amount > 0:
                            tranche_text = f"- Buy {quantity:.2f} shares at ${price_level:,.2f}, {status}, ${amount:,.0f}"
                        else:
                            tranche_text = f"- {status}"
                    elements.append(Paragraph(tranche_text, styles["Normal"]))

            elements.append(Spacer(1, 0.05 * inch))

            # Buy Quality for each support level
            s1_quality = pos.get("s1_quality", "N/A")
            s1_quality_note = pos.get("s1_quality_note", "")
            s2_quality = pos.get("s2_quality", "N/A")
            s2_quality_note = pos.get("s2_quality_note", "")
            s3_quality = pos.get("s3_quality", "N/A")
            s3_quality_note = pos.get("s3_quality_note", "")

            quality_text = f"""
            <b>Buy Quality S1:</b> {s1_quality} - {s1_quality_note}<br/>
            <b>Buy Quality S2:</b> {s2_quality} - {s2_quality_note}<br/>
            <b>Buy Quality S3:</b> {s3_quality} - {s3_quality_note}
            """
            elements.append(Paragraph(quality_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Volume Profile at the end
            volume_text = f"""
            <b>Volume Profile (60d):</b> POC ${pos.get('poc_60d', 0):,.2f} | VAL ${pos.get('val_60d', 0):,.2f} - VAH ${pos.get('vah_60d', 0):,.2f}
            """
            elements.append(Paragraph(volume_text, styles["Normal"]))

            elements.append(Spacer(1, 0.15 * inch))
    else:
        elements.append(Paragraph("BUY ACTIONS (0 positions)", section_style))
        elements.append(
            Paragraph("No buy opportunities at this time.", styles["Normal"])
        )
        elements.append(PageBreak())

    # === SECTION 3: SELL ACTIONS ===
    sell_positions = [
        p for p in portfolio_data["positions"] if p.get("action") == "SELL"
    ]

    if sell_positions:
        elements.append(PageBreak())
        elements.append(
            Paragraph(f"SELL ACTIONS ({len(sell_positions)} positions)", section_style)
        )
        elements.append(Spacer(1, 0.1 * inch))

        for pos in sell_positions:
            ticker = pos["ticker"]
            signal = pos["signal"]
            price = pos["current_price"]
            current_value = pos.get("current_value", 0)
            target_pct = pos.get("target_pct", 0) * 100
            overweight = current_value - (
                pos.get("gap_value", 0)
            )  # How much over target

            # Position header
            header_text = f"<b>{ticker}</b> - ${price:,.2f} ({signal})"
            elements.append(Paragraph(header_text, subsection_style))

            # Allocation info
            alloc_text = f"""
            <b>Allocation:</b> Target {target_pct:.0f}% | Current ${current_value:,.0f} | Overweight by ${abs(pos.get('gap_value', 0)):,.0f}
            """
            elements.append(Paragraph(alloc_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Show MA feasibility note before reduction plan to explain adjusted levels
            sell_note = pos.get("sell_feasibility_note", "")
            if sell_note and sell_note != "Clear path to all R-levels":
                note_text = f"<i><b>Note:</b> {sell_note} - Reduction plan uses adjusted levels</i>"
                elements.append(Paragraph(note_text, styles["Normal"]))
                elements.append(Spacer(1, 0.05 * inch))

            # Sell tranches (simplified - no quality notes repeated)
            tranches = pos.get("sell_tranches", [])
            if tranches:
                elements.append(Paragraph("<b>Reduction Plan:</b>", styles["Normal"]))
                total_reduction = 0
                # Handle both old (5 items) and new (7 items) tuple formats
                for tranche in tranches:
                    if len(tranche) == 7:
                        (
                            price_level,
                            amount,
                            pct,
                            level_name,
                            status,
                            quality,
                            quality_note,
                        ) = tranche
                    else:
                        price_level, amount, pct, level_name, status = tranche[:5]

                    quantity = (
                        amount / price_level if price_level and price_level > 0 else 0
                    )
                    if price_level is not None:
                        tranche_text = f"- {level_name} at ${price_level:,.2f}: Sell {quantity:.2f} shares ({pct*100:.0f}%) = ${amount:,.0f}"
                    else:
                        tranche_text = f"- {level_name}: Sell {quantity:.2f} shares ({pct*100:.0f}%) = ${amount:,.0f}"
                    elements.append(Paragraph(tranche_text, styles["Normal"]))
                    total_reduction += amount

                elements.append(Spacer(1, 0.05 * inch))

                keep_amount = current_value - total_reduction
                summary_text = f"<b>Result:</b> Reduce ${total_reduction:,.0f}, Keep ${keep_amount:,.0f}"
                elements.append(Paragraph(summary_text, styles["Normal"]))

            elements.append(Spacer(1, 0.05 * inch))

            # Sell Quality for each resistance level
            r1_quality = pos.get("r1_quality", "N/A")
            r1_quality_note = pos.get("r1_quality_note", "")
            r2_quality = pos.get("r2_quality", "N/A")
            r2_quality_note = pos.get("r2_quality_note", "")
            r3_quality = pos.get("r3_quality", "N/A")
            r3_quality_note = pos.get("r3_quality_note", "")

            quality_text = f"""
            <b>Sell Quality R1:</b> {r1_quality} - {r1_quality_note}<br/>
            <b>Sell Quality R2:</b> {r2_quality} - {r2_quality_note}<br/>
            <b>Sell Quality R3:</b> {r3_quality} - {r3_quality_note}
            """
            elements.append(Paragraph(quality_text, styles["Normal"]))
            elements.append(Spacer(1, 0.05 * inch))

            # Volume Profile at the end
            volume_text = f"""
            <b>Volume Profile (60d):</b> POC ${pos.get('poc_60d', 0):,.2f} | VAL ${pos.get('val_60d', 0):,.2f} - VAH ${pos.get('vah_60d', 0):,.2f}
            """
            elements.append(Paragraph(volume_text, styles["Normal"]))

            elements.append(Spacer(1, 0.15 * inch))
    else:
        elements.append(Paragraph("SELL ACTIONS (0 positions)", section_style))
        elements.append(Paragraph("No reduction actions needed.", styles["Normal"]))

    # === SECTION 4: HOLD POSITIONS ===
    hold_positions = [
        p for p in portfolio_data["positions"] if p.get("action") in ["HOLD", "WAIT"]
    ]

    if hold_positions:
        elements.append(PageBreak())
        elements.append(
            Paragraph(
                f"HOLD POSITIONS ({len(hold_positions)} positions)", section_style
            )
        )
        elements.append(Spacer(1, 0.1 * inch))

        for pos in hold_positions:
            ticker = pos["ticker"]
            signal = pos["signal"]
            price = pos["current_price"]
            action_reason = (
                "At target allocation"
                if pos.get("gap_value", 0) == 0
                else "Waiting for signal/setup"
            )

            hold_text = (
                f"- <b>{ticker}</b> - ${price:,.2f} ({signal}) - {action_reason}"
            )
            elements.append(Paragraph(hold_text, styles["Normal"]))

    # Build PDF
    doc.build(elements)
    return output_path


def create_portfolio_tracker_excel(portfolio_data, output_path):
    """
    Create Portfolio Tracker Excel with 3 tabs:
    1. Trade Log (pending/executed trades)
    2. Current Positions (allocation tracker)
    3. Technical Levels (S/R + MAs)

    Args:
        portfolio_data: dict with 'positions' list and 'portfolio_total'
        output_path: Path to save Excel file

    Returns:
        Path to created file
    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Define styles
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    # === TAB 1: TRADE LOG ===
    ws_trade_log = wb.create_sheet("Trade Log")

    # Add portfolio summary at the top
    summary = portfolio_data.get("summary", {})
    total_value = summary.get("total_portfolio_value", 0)
    current_holdings = portfolio_data["portfolio_total"]
    cash_available = summary.get("cash_available", total_value - current_holdings)
    cash_pct = (cash_available / total_value * 100) if total_value > 0 else 0

    ws_trade_log.cell(1, 1, "TRADE LOG").font = Font(bold=True, size=14)
    ws_trade_log.cell(
        2,
        1,
        "Trade history from Portfolio Manager UI. Use http://localhost:8501 to add/edit trades (run: .\\scripts\\start_ui.ps1)",
    )
    ws_trade_log.cell(2, 1).font = Font(italic=True, color="666666")

    # Trade log headers at row 4
    trade_log_headers = [
        "Date",
        "Ticker",
        "Action",
        "Price",
        "Quantity",
        "Amount",
        "Notes",
    ]

    for col_num, header in enumerate(trade_log_headers, 1):
        cell = ws_trade_log.cell(4, col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Load and populate from transactions.csv
    from pathlib import Path

    transactions_file = Path(__file__).parent.parent / "data" / "transactions.csv"

    if transactions_file.exists():
        try:
            transactions_df = pd.read_csv(transactions_file)
            # Filter out empty rows
            transactions_df = transactions_df.dropna(
                subset=["date", "ticker"], how="all"
            )

            # Sort by date descending (most recent first)
            if not transactions_df.empty and "date" in transactions_df.columns:
                transactions_df["date"] = pd.to_datetime(
                    transactions_df["date"], errors="coerce"
                )
                transactions_df = transactions_df.sort_values("date", ascending=False)

                # Populate trade log
                row = 5
                for _, txn in transactions_df.iterrows():
                    ws_trade_log.cell(
                        row,
                        1,
                        (
                            txn.get("date").strftime("%Y-%m-%d")
                            if pd.notna(txn.get("date"))
                            else ""
                        ),
                    )
                    ws_trade_log.cell(row, 2, txn.get("ticker", ""))
                    ws_trade_log.cell(
                        row, 3, str(txn.get("type", "")).upper()
                    )  # BUY/SELL
                    ws_trade_log.cell(
                        row,
                        4,
                        (
                            f"${txn.get('price', 0):,.2f}"
                            if pd.notna(txn.get("price"))
                            else ""
                        ),
                    )
                    ws_trade_log.cell(
                        row,
                        5,
                        (
                            f"{txn.get('quantity', 0):,.4f}"
                            if pd.notna(txn.get("quantity"))
                            else ""
                        ),
                    )
                    ws_trade_log.cell(
                        row,
                        6,
                        (
                            f"${txn.get('total_value', 0):,.2f}"
                            if pd.notna(txn.get("total_value"))
                            else ""
                        ),
                    )
                    ws_trade_log.cell(row, 7, txn.get("notes", ""))
                    row += 1
        except Exception as e:
            logger.warning(f"Could not load transactions.csv: {e}")

    # Auto-size columns
    for col in ws_trade_log.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_trade_log.column_dimensions[col[0].column_letter].width = max(
            max_length + 2, 12
        )

    # === TAB 2: ACTION PLAN ===
    ws_actions = wb.create_sheet("Action Plan")

    ws_actions.cell(1, 1, "PORTFOLIO SUMMARY").font = Font(bold=True, size=14)
    ws_actions.cell(2, 1, "Total Portfolio Value:")
    ws_actions.cell(2, 2, f"${total_value:,.0f}")
    ws_actions.cell(3, 1, "Current Holdings:")
    ws_actions.cell(3, 2, f"${current_holdings:,.0f}")
    ws_actions.cell(4, 1, "Cash Available:")
    ws_actions.cell(4, 2, f"${cash_available:,.0f} (~{cash_pct:.0f}% dry powder)")
    ws_actions.cell(4, 2).font = Font(bold=True)

    # Action plan starts at row 6
    action_headers = [
        "Date",
        "Ticker",
        "Action",
        "Signal",
        "Price",
        "Quantity",
        "Amount",
        "Level",
        "Quality",
        "Quality Note",
        "Status",
    ]

    for col_num, header in enumerate(action_headers, 1):
        cell = ws_actions.cell(6, col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add recommended actions
    row = 7
    today = datetime.now().strftime("%Y-%m-%d")

    for pos in portfolio_data["positions"]:
        ticker = pos["ticker"]
        signal = pos["signal"]
        tradeable_quantity = pos.get("tradeable_quantity", 0)
        action = pos.get("action", "HOLD")

        # Only show buy tranches for stocks with BUY action and FULL HOLD + ADD signal
        if signal == "FULL HOLD + ADD" and action == "BUY":
            # Reverse to show S1 first (shallowest/most likely), then S2, then S3
            for tranche in reversed(pos.get("buy_tranches", [])):
                # Handle both old (4 items) and new (6 items) tuple formats
                if len(tranche) == 6:
                    price_level, amount, level_name, status, quality, quality_note = (
                        tranche
                    )
                else:
                    price_level, amount, level_name, status = tranche[:4]
                    quality, quality_note = "N/A", ""

                # Calculate quantity
                quantity = (
                    amount / price_level
                    if price_level and price_level > 0 and amount > 0
                    else 0
                )

                ws_actions.cell(row, 1, today)
                ws_actions.cell(row, 2, ticker)
                ws_actions.cell(row, 3, "BUY")
                ws_actions.cell(row, 4, signal)
                ws_actions.cell(
                    row,
                    5,
                    f"${price_level:,.2f}" if price_level and price_level > 0 else "-",
                )
                ws_actions.cell(row, 6, f"{quantity:.2f}" if quantity > 0 else "-")
                ws_actions.cell(row, 7, f"${amount:,.2f}" if amount > 0 else "-")
                ws_actions.cell(row, 8, level_name)
                ws_actions.cell(row, 9, quality)
                ws_actions.cell(row, 10, quality_note)
                ws_actions.cell(row, 11, f"{status}")
                row += 1

        # Only add sell tranches if we actually have a position to sell
        if tradeable_quantity > 0:
            for tranche in pos.get("sell_tranches", []):
                # Handle both old (5 items) and new (7 items) tuple formats
                if len(tranche) == 7:
                    (
                        price_level,
                        amount,
                        pct,
                        level_name,
                        status,
                        quality,
                        quality_note,
                    ) = tranche
                else:
                    price_level, amount, pct, level_name, status = tranche[:5]
                    quality, quality_note = "N/A", ""

                # Calculate quantity
                quantity = amount / price_level if price_level > 0 else 0

                ws_actions.cell(row, 1, today)
                ws_actions.cell(row, 2, ticker)
                ws_actions.cell(row, 3, "SELL")
                ws_actions.cell(row, 4, signal)
                ws_actions.cell(row, 5, f"${price_level:,.2f}")
                ws_actions.cell(row, 6, f"{quantity:.2f}")
                ws_actions.cell(row, 7, f"${amount:,.2f}")
                ws_actions.cell(row, 8, level_name)
                ws_actions.cell(row, 9, quality)
                ws_actions.cell(row, 10, quality_note)
                ws_actions.cell(row, 11, status)
                row += 1

    # Auto-size columns
    for col in ws_actions.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_actions.column_dimensions[col[0].column_letter].width = max_length + 2

    # === TAB 3: CURRENT POSITIONS ===
    ws_positions = wb.create_sheet("Current Positions")
    pos_headers = [
        "Ticker",
        "Signal",
        "Price",
        "Target%",
        "Current$",
        "Target$",
        "Gap$",
        "Action",
        "Quality",
        "Quality Note",
    ]

    for col_num, header in enumerate(pos_headers, 1):
        cell = ws_positions.cell(1, col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add positions
    row = 2
    for pos in portfolio_data["positions"]:
        ws_positions.cell(row, 1, pos["ticker"])
        ws_positions.cell(row, 2, pos["signal"])
        ws_positions.cell(row, 3, pos["current_price"])
        ws_positions.cell(
            row, 4, int(pos.get("target_pct", 0) * 100)
        )  # Format as whole number
        ws_positions.cell(row, 5, pos.get("current_value", 0))
        ws_positions.cell(
            row, 6, pos.get("target_pct", 0) * portfolio_data["portfolio_total"]
        )
        ws_positions.cell(row, 7, pos.get("gap_value", 0))

        # Determine action and quality based on signal
        action = pos.get("action", "HOLD")
        ws_positions.cell(row, 8, action)

        # Show sell quality for SELL actions, buy quality otherwise
        if action == "SELL":
            quality = pos.get("r1_quality", "N/A")
            quality_note = pos.get("r1_quality_note", "")
        else:
            quality = pos.get("buy_quality", "N/A")
            quality_note = pos.get("buy_quality_note", "")

        ws_positions.cell(row, 9, quality)
        ws_positions.cell(row, 10, quality_note)
        row += 1

    # Auto-size columns
    for col in ws_positions.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_positions.column_dimensions[col[0].column_letter].width = max_length + 2

    # === TAB 4: TECHNICAL LEVELS ===
    ws_tech = wb.create_sheet("Technical Levels")
    tech_headers = [
        "Ticker",
        "Signal",
        "Price",
        "RSI",
        "BB_Upper",
        "BB_Middle",
        "BB_Lower",
        "D50",
        "D100",
        "D200",
        "POC_60d",
        "VAH_60d",
        "VAL_60d",
        "HVN_Above",
        "HVN_Below",
        "LVN_Above",
        "LVN_Below",
        "S1",
        "S2",
        "S3",
        "R1",
        "R2",
        "R3",
    ]

    for col_num, header in enumerate(tech_headers, 1):
        cell = ws_tech.cell(1, col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add technical data
    row = 2
    for pos in portfolio_data["positions"]:
        ws_tech.cell(row, 1, pos["ticker"])
        ws_tech.cell(row, 2, pos["signal"])
        ws_tech.cell(row, 3, pos["current_price"])
        # RSI and Bollinger Bands
        ws_tech.cell(row, 4, pos.get("rsi"))
        ws_tech.cell(row, 5, pos.get("bb_upper"))
        ws_tech.cell(row, 6, pos.get("bb_middle"))
        ws_tech.cell(row, 7, pos.get("bb_lower"))
        # Moving Averages
        ws_tech.cell(row, 8, pos.get("d50"))
        ws_tech.cell(row, 9, pos.get("d100"))
        ws_tech.cell(row, 10, pos.get("d200"))
        # VRVP 60-day data
        ws_tech.cell(row, 11, pos.get("poc_60d"))
        ws_tech.cell(row, 12, pos.get("vah_60d"))
        ws_tech.cell(row, 13, pos.get("val_60d"))
        ws_tech.cell(row, 14, pos.get("hvn_above_60d"))
        ws_tech.cell(row, 15, pos.get("hvn_below_60d"))
        ws_tech.cell(row, 16, pos.get("lvn_above_60d"))
        ws_tech.cell(row, 17, pos.get("lvn_below_60d"))
        # Support/Resistance
        ws_tech.cell(row, 18, pos.get("s1"))
        ws_tech.cell(row, 19, pos.get("s2"))
        ws_tech.cell(row, 20, pos.get("s3"))
        ws_tech.cell(row, 21, pos.get("r1"))
        ws_tech.cell(row, 22, pos.get("r2"))
        ws_tech.cell(row, 23, pos.get("r3"))
        row += 1

    # Auto-size columns
    for col in ws_tech.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_tech.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save workbook
    wb.save(output_path)
    return output_path


def cleanup_old_reports(results_dir, max_files=3, archive_retention_days=90):
    """
    Keep only the most recent portfolio reports
    Move older ones to archive, delete archive files older than archive_retention_days

    Args:
        results_dir: Directory containing portfolio results
        max_files: Number of most recent files to keep
        archive_retention_days: Days to keep files in archive before deletion
    """
    from datetime import datetime, timedelta
    import time

    archive_dir = results_dir / "archive"
    archive_dir.mkdir(exist_ok=True)

    total_archived = 0

    # Get all PDF and Excel files
    pdf_files = sorted(
        results_dir.glob("portfolio_playbook_*.pdf"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    xlsx_files = sorted(
        results_dir.glob("portfolio_tracker_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    # Move files beyond max_files to archive
    skipped_files = []

    for old_file in pdf_files[max_files:]:
        archive_path = archive_dir / old_file.name
        try:
            old_file.rename(archive_path)
            print(f"  📦 Archived: {old_file.name}")
            total_archived += 1
        except PermissionError:
            skipped_files.append(old_file.name)
        except Exception as e:
            logger.warning(f"Could not archive {old_file.name}: {e}")
            print(f"  ⚠️  Could not archive {old_file.name}: {e}")

    for old_file in xlsx_files[max_files:]:
        archive_path = archive_dir / old_file.name
        try:
            old_file.rename(archive_path)
            print(f"  📦 Archived: {old_file.name}")
            total_archived += 1
        except PermissionError:
            skipped_files.append(old_file.name)
        except Exception as e:
            logger.warning(f"Could not archive {old_file.name}: {e}")
            print(f"  ⚠️  Could not archive {old_file.name}: {e}")

    if total_archived > 0:
        print(f"  ✅ Archived {total_archived} file(s), kept {max_files} most recent")
    else:
        print(f"  ✅ No files to archive (only {max_files} or fewer exist)")

    if skipped_files:
        print(
            f"  ⚠️  Skipped {len(skipped_files)} file(s) - close Excel/PDF viewer and run again:"
        )
        for filename in skipped_files:
            print(f"     • {filename}")

    # Delete archive files older than retention period
    cutoff_time = time.time() - (archive_retention_days * 86400)
    deleted_count = 0

    for archive_file in archive_dir.glob("*"):
        if archive_file.is_file() and archive_file.stat().st_mtime < cutoff_time:
            try:
                archive_file.unlink()
                deleted_count += 1
            except Exception as e:
                logger.warning(
                    f"Could not delete archive file {archive_file.name}: {e}"
                )
                print(f"  ⚠️  Could not delete {archive_file.name}: {e}")

    if deleted_count > 0:
        print(
            f"  🗑️  Deleted {deleted_count} archive file(s) older than {archive_retention_days} days"
        )
